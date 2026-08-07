"""
ORCMM - Lector de las fuentes que ya no caben en el Excel.

Dos hojas del layout pasaron del millón de filas que aguanta una hoja de
Excel y se entregan aparte, como CSV: TABLEAU_INV_TIENDA (2.7 millones de
filas, repartidas en 6 archivos) y TABLEAU_VENTAS. El resto del layout sigue
viniendo en el .xlsx.

Este módulo lee esos CSV y entrega filas con la MISMA forma que devuelve
leer_hoja() del pipeline: un dict por fila, con los nombres canónicos del
spec como llaves. Así el resto del programa no tiene que saber de dónde vino
cada hoja.

Tres cosas que los archivos reales obligan a resolver aquí:

  - No son CSV de coma. Salen de Tableau en UTF-16 LE con BOM y separados
    por TAB. El encoding y el delimitador se detectan, no se asumen.
  - Los encabezados vienen con nombre de negocio ("Código de Barras" por
    sku), y la columna de la métrica viene SIN ENCABEZADO. Se traducen con
    la misma idea de ALIAS_ENCABEZADOS del spec, y la columna sin nombre se
    resuelve por posición.
  - Las fechas vienen como texto y en dos idiomas: "February 16, 2026" en
    inventario, "1 de Marzo de 2026" en ventas.

Se lee en streaming, una fila a la vez: los 6 archivos de inventario suman
222 MB y la máquina donde corre esto tiene 512 MB. Por eso leer_csv() es un
generador y acepta un filtro de llaves, para que quien llame se quede sólo
con lo que va a usar.

    python orcmm_fuentes_csv.py datos_reales_v5/TABLEAU_VENTAS.csv

corre el módulo como inspector: dice qué detectó y qué leería.
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
import unicodedata
from datetime import date, datetime
from functools import lru_cache
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Optional, Set, Tuple

from orcmm_layout_spec import (FILA_DATOS, FILA_ENCABEZADO, HOJAS,
                               normalizar_encabezado)

# Hojas del layout que hoy se entregan como CSV. El resto sigue en el .xlsx.
HOJAS_CSV = ("TABLEAU_INV_TIENDA", "TABLEAU_VENTAS")


# ---------------------------------------------------------------------------
# Encabezados
#
# Mismo criterio que ALIAS_ENCABEZADOS en el spec: el equipo entrega la
# columna con el nombre de negocio y aquí se traduce al nombre del campo, en
# vez de romper la lectura. La llave se compara en minúsculas y con los
# espacios colapsados.
# ---------------------------------------------------------------------------

ALIAS_CSV = {
    "TABLEAU_INV_TIENDA": {
        "codigo de barras": "sku",
        "sku": "sku",
        "tienda no": "tienda",
        "tienda no.": "tienda",
        "day of fecha": "fecha",
        "dia en texto": "fecha",
        "existencia": "existencia_piezas",
    },
    "TABLEAU_VENTAS": {
        "codigo de barras": "sku",
        "sku": "sku",
        "tienda no": "tienda",
        "tienda no.": "tienda",
        "dia en texto": "fecha",
        "day of fecha": "fecha",
    },
}

# A qué campo corresponde la columna que viene SIN ENCABEZADO.
#
# El export de Tableau deja el nombre de la medida en blanco, así que no hay
# forma de deducirlo del archivo: se declara aquí. Confirmado con La Comer el
# 2026-08-06: en ventas es el importe, no las unidades.
COLUMNA_SIN_NOMBRE = {
    "TABLEAU_INV_TIENDA": "existencia_piezas",
    "TABLEAU_VENTAS": "importe_venta",
}

# Claves que marcan la fila de totales que Tableau agrega al final del export.
TOTALES = {"total", "grand total", "gran total", "suma", "sum"}


def _sin_acentos(s: str) -> str:
    """Compara encabezados ignorando acentos.

    Un export mal decodificado en algún paso intermedio llega como
    "C?digo de Barras" o "Codigo de Barras". La tabla de alias se escribe una
    sola vez, sin acentos, y aquí se normaliza lo que venga.
    """
    return "".join(c for c in unicodedata.normalize("NFD", s)
                   if unicodedata.category(c) != "Mn")


def normalizar_encabezado_csv(hoja: str, texto) -> Optional[str]:
    """Nombre canónico del campo a partir del encabezado tal como vino."""
    if texto is None:
        return None
    limpio = " ".join(str(texto).split())
    if not limpio:
        return None
    llave = _sin_acentos(limpio).lower()
    return ALIAS_CSV.get(hoja, {}).get(llave, limpio)


def hoja_de_archivo(nombre: str) -> Optional[str]:
    """A qué hoja del layout pertenece un CSV, según su nombre de archivo.

    Los cortes por fecha llegan numerados (TABLEAU_INV_TIENDA_1.csv ...
    _6.csv) y son la misma tabla partida, así que se resuelven por prefijo:
    agregar un séptimo archivo no cuesta nada.
    """
    base = Path(nombre).stem.strip().upper()
    candidatas = [h for h in HOJAS_CSV if base == h or base.startswith(h + "_")]
    # La más larga gana, por si dos hojas comparten prefijo en el futuro.
    return max(candidatas, key=len) if candidatas else None


# ---------------------------------------------------------------------------
# Encoding y delimitador
# ---------------------------------------------------------------------------

# UTF-32 va primero a propósito: su BOM empieza con el de UTF-16 LE y si se
# revisara al revés, un archivo UTF-32 se leería como UTF-16.
BOMS = [
    (b"\xff\xfe\x00\x00", "utf-32"),
    (b"\x00\x00\xfe\xff", "utf-32"),
    (b"\xff\xfe", "utf-16"),
    (b"\xfe\xff", "utf-16"),
    (b"\xef\xbb\xbf", "utf-8-sig"),
]

DELIMITADORES = ["\t", ";", ",", "|"]


def detectar_encoding(ruta: Path) -> str:
    """Encoding del archivo, por BOM y con verificación.

    Sin BOM se prueba UTF-8 y, si no decodifica, se cae a cp1252: es lo que
    sale de un Excel en español guardado como CSV, y nunca falla al
    decodificar, así que va de último.
    """
    with ruta.open("rb") as f:
        cabeza = f.read(4)
    for bom, enc in BOMS:
        if cabeza.startswith(bom):
            return enc

    with ruta.open("rb") as f:
        muestra = f.read(64 * 1024)
    try:
        muestra.decode("utf-8")
        return "utf-8"
    except UnicodeDecodeError:
        return "cp1252"


def detectar_delimitador(linea: str) -> str:
    """El separador más frecuente en la línea de encabezados.

    No se usa csv.Sniffer: con una fila de importes como "201,897.00" se
    confunde y elige la coma. Contar sobre el encabezado, que no trae
    números, es más confiable que adivinar sobre los datos.
    """
    conteo = {d: linea.count(d) for d in DELIMITADORES}
    mejor = max(conteo, key=lambda d: conteo[d])
    return mejor if conteo[mejor] else "\t"


# ---------------------------------------------------------------------------
# Conversión de valores
# ---------------------------------------------------------------------------

MESES = {}
for _i, (_en, _es) in enumerate([
    ("january", "enero"), ("february", "febrero"), ("march", "marzo"),
    ("april", "abril"), ("may", "mayo"), ("june", "junio"),
    ("july", "julio"), ("august", "agosto"), ("september", "septiembre"),
    ("october", "octubre"), ("november", "noviembre"), ("december", "diciembre"),
], 1):
    MESES[_en] = _i
    MESES[_es] = _i
    MESES[_en[:3]] = _i
    MESES[_es[:3]] = _i

# "February 16, 2026" / "Feb 16 2026"
_EN = re.compile(r"^([a-záéíóú]+)\.?\s+(\d{1,2})\s*,?\s+(\d{4})$", re.I)
# "1 de Marzo de 2026" / "1 marzo 2026"
_ES = re.compile(r"^(\d{1,2})\s+(?:de\s+)?([a-záéíóú]+)\.?\s+(?:de\s+)?(\d{4})$", re.I)
# "16/02/2026" o "16-02-2026" — día primero, que es como se captura en México
_NUM = re.compile(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$")


@lru_cache(maxsize=4096)
def _fecha_de_texto(s: str) -> Optional[date]:
    """El parseo en sí, memorizado.

    Un mes de inventario son 2.7 millones de filas y 44 fechas distintas: sin
    caché se corren los mismos tres regex millones de veces para volver a
    obtener el mismo día. El tope de 4096 entradas es holgado para cualquier
    periodo de análisis y acota la memoria si llegara basura variada.
    """
    if not s or s.lower() in TOTALES:
        return None

    try:                                   # ISO, con o sin hora
        return date.fromisoformat(s[:10])
    except ValueError:
        pass

    m = _EN.match(s)
    if m and m.group(1).lower() in MESES:
        return date(int(m.group(3)), MESES[m.group(1).lower()], int(m.group(2)))

    m = _ES.match(s)
    if m and m.group(2).lower() in MESES:
        return date(int(m.group(3)), MESES[m.group(2).lower()], int(m.group(1)))

    m = _NUM.match(s)
    if m:
        return date(int(m.group(3)), int(m.group(2)), int(m.group(1)))

    return None


def fecha_csv(v) -> Optional[date]:
    """Fecha a partir del texto tal como lo escribe Tableau.

    Se aceptan los dos idiomas a propósito: el mismo layout entrega
    inventario en inglés y ventas en español, y no se le va a pedir al equipo
    que unifique el formato de un export automático.
    """
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return _fecha_de_texto(" ".join(str(v).split()))


def numero_csv(v) -> Optional[float]:
    """Número a partir del texto, tolerando separadores de miles.

    Se decide cuál de los dos separadores es el decimal por cuál aparece de
    último: "201,897.00" y "201.897,00" son el mismo número escrito en dos
    convenciones, y el export puede traer cualquiera de las dos según cómo
    esté configurada la máquina que lo generó.
    """
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)

    s = str(v).strip().replace(" ", "").replace("\xa0", "")
    if not s or s.lower() in TOTALES:
        return None

    negativo = s.startswith("(") and s.endswith(")")     # contabilidad: (1.00)
    if negativo:
        s = s[1:-1]
    s = s.replace("$", "").replace("%", "")

    if "," in s and "." in s:
        decimal = "," if s.rfind(",") > s.rfind(".") else "."
        s = s.replace("." if decimal == "," else ",", "").replace(decimal, ".")
    elif "," in s:
        # Sólo coma: es decimal salvo que agrupe de tres en tres ("201,897").
        entero, _, resto = s.rpartition(",")
        s = s.replace(",", "") if len(resto) == 3 and entero else s.replace(",", ".")

    try:
        n = float(s)
    except ValueError:
        return None
    return -n if negativo else n


def texto_csv(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


# Cómo convertir cada campo. Lo que no esté aquí se queda como texto.
CONVERSORES = {
    "fecha": fecha_csv,
    "existencia_piezas": numero_csv,
    "existencia_minima_dia": numero_csv,
    "unidades_vendidas": numero_csv,
    "importe_venta": numero_csv,
    "venta_perdida_estimada": numero_csv,
}


# ---------------------------------------------------------------------------
# Lectura
# ---------------------------------------------------------------------------

class ReporteCSV:
    """Qué se leyó y qué se descartó, para poder explicarlo después.

    El pipeline reporta por hoja cuántas filas entraron y en qué rango de
    fechas. Como aquí se descarta al vuelo, esos conteos hay que llevarlos
    durante el recorrido: si se contaran sobre lo que quedó en memoria, la
    hoja "Cobertura y fuentes" reportaría el filtro y no el archivo.
    """

    def __init__(self, hoja: str):
        self.hoja = hoja
        self.archivos: List[str] = []
        self.filas_leidas = 0
        self.filas_devueltas = 0
        self.sin_fecha = 0
        self.totales_descartados = 0
        self.desde: Optional[date] = None
        self.hasta: Optional[date] = None
        self.tiendas: Set[str] = set()
        self.advertencias: List[str] = []

    def avisar(self, msg: str) -> None:
        """Una advertencia por motivo, no por archivo.

        Los 6 cortes de inventario comparten encabezado: sin esto, cada
        observación sobre las columnas saldría repetida seis veces.
        """
        if msg not in self.advertencias:
            self.advertencias.append(msg)

    def ver_fecha(self, d: date) -> None:
        if self.desde is None or d < self.desde:
            self.desde = d
        if self.hasta is None or d > self.hasta:
            self.hasta = d

    def __repr__(self) -> str:
        return (f"<{self.hoja}: {self.filas_leidas} leídas, "
                f"{self.filas_devueltas} devueltas, {self.desde}..{self.hasta}>")


def _resolver_encabezados(hoja: str, crudos: List[str],
                          rep: ReporteCSV) -> List[Optional[str]]:
    """Nombre canónico de cada columna, incluida la que viene sin encabezado."""
    campos = [c[0] for c in HOJAS[hoja]["campos"]]
    nombres = [normalizar_encabezado_csv(hoja, c) for c in crudos]

    sin_nombre = [i for i, n in enumerate(nombres) if not n]
    if sin_nombre:
        supuesto = COLUMNA_SIN_NOMBRE.get(hoja)
        # Sólo la primera columna sin encabezado se resuelve por posición: si
        # vienen dos, adivinar cuál es cuál sería inventar el dato.
        if supuesto and len(sin_nombre) == 1:
            nombres[sin_nombre[0]] = supuesto
            rep.avisar(
                f"{hoja}: la columna {sin_nombre[0] + 1} viene sin encabezado y se "
                f"está leyendo como '{supuesto}'. Pedir que el export la nombre.")
        else:
            rep.avisar(
                f"{hoja}: {len(sin_nombre)} columnas sin encabezado que no se "
                f"pueden asignar a un campo. Se ignoran.")

    for crudo, canonico in zip(crudos, nombres):
        if crudo and canonico and crudo != canonico and canonico in campos:
            rep.avisar(
                f"{hoja}: la columna '{crudo}' se está leyendo como '{canonico}'.")

    desconocidas = [n for n in nombres if n and n not in campos]
    if desconocidas:
        rep.avisar(
            f"{hoja}: columnas que el spec no reconoce -> {', '.join(desconocidas)}.")

    obligatorios = [c[0] for c in HOJAS[hoja]["campos"] if c[2]]
    faltan = [c for c in obligatorios if c not in nombres]
    if faltan:
        rep.avisar(
            f"{hoja}: faltan columnas obligatorias -> {', '.join(faltan)}.")

    return nombres


def leer_csv(rutas: Iterable[Path], hoja: str,
             llaves: Optional[Set[Tuple]] = None,
             campos_llave: Tuple[str, ...] = ("sku", "tienda", "fecha"),
             rep: Optional[ReporteCSV] = None) -> Iterator[dict]:
    """Filas de una hoja entregada como CSV, ya convertidas y en streaming.

    `rutas` puede traer varios archivos: los cortes por fecha de una misma
    hoja se concatenan y salen como una sola tabla.

    `llaves`, si se pasa, deja pasar únicamente las filas cuya llave natural
    esté en el conjunto. Es lo que hace viable leer 2.7 millones de filas de
    inventario para quedarse con los 30 mil días que el análisis va a mirar:
    lo que no está en el conjunto se descarta sin construir el dict.
    """
    rep = rep if rep is not None else ReporteCSV(hoja)

    for ruta in rutas:
        ruta = Path(ruta)
        rep.archivos.append(ruta.name)
        enc = detectar_encoding(ruta)

        with ruta.open("r", encoding=enc, newline="", errors="replace") as f:
            primera = f.readline()
            if not primera.strip():
                rep.avisar(f"{ruta.name}: archivo vacío.")
                continue
            delim = detectar_delimitador(primera)

            crudos = next(csv.reader([primera], delimiter=delim), [])
            nombres = _resolver_encabezados(hoja, crudos, rep)
            conversores = [CONVERSORES.get(n, texto_csv) if n else None
                           for n in nombres]
            # Posición de cada campo de la llave, para filtrar sin armar el
            # dict de las filas que no interesan.
            try:
                pos_llave = [nombres.index(c) for c in campos_llave]
            except ValueError:
                pos_llave = []

            for fila in csv.reader(f, delimiter=delim):
                if not fila or all(not (c or "").strip() for c in fila):
                    continue
                rep.filas_leidas += 1

                if (fila[0] or "").strip().lower() in TOTALES:
                    rep.totales_descartados += 1
                    continue

                if pos_llave:
                    try:
                        llave = tuple(conversores[i](fila[i]) for i in pos_llave)
                    except IndexError:
                        continue
                    if llave[-1] is None:          # la fecha no se pudo leer
                        rep.sin_fecha += 1
                        continue
                    if llaves is not None and llave not in llaves:
                        continue

                registro = {}
                for i, nombre in enumerate(nombres):
                    if not nombre or i >= len(fila):
                        continue
                    registro[nombre] = conversores[i](fila[i])

                d = registro.get("fecha")
                if isinstance(d, date):
                    rep.ver_fecha(d)
                t = registro.get("tienda")
                if t:
                    rep.tiendas.add(t)

                rep.filas_devueltas += 1
                yield registro

    return


def agrupar_por_hoja(rutas: Iterable[Path]) -> Tuple[Dict[str, List[Path]], List[Path]]:
    """Reparte los CSV entre las hojas del layout, por nombre de archivo.

    Los que no correspondan a ninguna hoja se regresan aparte: ignorarlos en
    silencio es la forma más fácil de correr un análisis al que le falta una
    fuente sin enterarse.
    """
    por_hoja: Dict[str, List[Path]] = {}
    sueltos: List[Path] = []
    for ruta in sorted(Path(r) for r in rutas):
        hoja = hoja_de_archivo(ruta.name)
        if hoja:
            por_hoja.setdefault(hoja, []).append(ruta)
        else:
            sueltos.append(ruta)
    return por_hoja, sueltos


def llaves_con_faltante_ws(ws, umbral_osa: float = 100.0) -> Set[Tuple[str, str, date]]:
    """Los días que el análisis va a mirar, leídos de la hoja BOPS_OSA.

    Es el conjunto con el que se filtra el CSV de inventario: el inventario de
    un día con el anaquel lleno no cambia ninguna conclusión, así que no vale
    la pena ni construirle el dict. Es también la referencia contra la que se
    cruzan las fuentes en la validación.

    OSA viene binario por diseño de BOPS (1 = producto visible ese día,
    0 = no visible) aunque el layout lo declare de 0 a 100, así que se
    normaliza multiplicando por 100 — mismo criterio que _osa_pct() en el
    pipeline.

    Las columnas se localizan por encabezado y no por posición: la hoja ganó
    una columna en el V5 y las posiciones se recorrieron.
    """
    llaves: Set[Tuple[str, str, date]] = set()
    encabezados = [normalizar_encabezado("BOPS_OSA", c.value)
                   for c in ws[FILA_ENCABEZADO]]
    try:
        i_sku, i_tda, i_fec, i_osa = (encabezados.index(c) for c in
                                      ("sku", "tienda", "fecha", "osa_pct"))
    except ValueError:
        return llaves

    tope = max(i_sku, i_tda, i_fec, i_osa)
    for fila in ws.iter_rows(min_row=FILA_DATOS, values_only=True):
        if len(fila) <= tope or fila[i_sku] is None:
            continue
        osa = numero_csv(fila[i_osa])
        if osa is None or osa * 100 >= umbral_osa:
            continue
        f = fecha_csv(fila[i_fec])
        if f:
            llaves.add((texto_csv(fila[i_sku]), texto_csv(fila[i_tda]), f))
    return llaves


def llaves_con_faltante(xlsx: Path, umbral_osa: float = 100.0) -> Set[Tuple[str, str, date]]:
    """Igual que llaves_con_faltante_ws, abriendo el archivo."""
    import openpyxl                       # sólo lo necesitan el inspector y el validador

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    try:
        if "BOPS_OSA" not in wb.sheetnames:
            return set()
        return llaves_con_faltante_ws(wb["BOPS_OSA"], umbral_osa)
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Inspector de línea de comandos
# ---------------------------------------------------------------------------

def main() -> int:
    p = argparse.ArgumentParser(
        description="Inspecciona un CSV de fuente: qué detecta y qué leería.")
    p.add_argument("archivos", nargs="+", type=Path)
    p.add_argument("--hoja", help="Forzar la hoja, si el nombre del archivo no la dice.")
    p.add_argument("--filas", type=int, default=3, help="Cuántas filas de muestra.")
    p.add_argument("--contra", type=Path, metavar="LAYOUT.xlsx",
                   help="Cruzar contra los días con faltante de BOPS_OSA y "
                        "reportar qué porcentaje quedaría cubierto.")
    args = p.parse_args()

    llaves = None
    if args.contra:
        llaves = llaves_con_faltante(args.contra)
        print(f"{args.contra.name}: {len(llaves):,} días con faltante en BOPS_OSA")
        tiendas = sorted({k[1] for k in llaves})
        print(f"  tiendas {tiendas[:10]}")

    por_hoja, sueltos = agrupar_por_hoja(args.archivos)
    if args.hoja:
        por_hoja = {args.hoja: [Path(a) for a in args.archivos]}
        sueltos = []

    for ruta in sueltos:
        print(f"  ?  {ruta.name}: no corresponde a ninguna hoja del layout "
              f"({', '.join(HOJAS_CSV)}).")

    for hoja, rutas in por_hoja.items():
        rep = ReporteCSV(hoja)
        print(f"\n=== {hoja} ===")
        for ruta in rutas:
            print(f"  {ruta.name}  ({ruta.stat().st_size / 1e6:.1f} MB, "
                  f"{detectar_encoding(ruta)})")

        muestra = []
        cubiertas = set()
        for fila in leer_csv(rutas, hoja, llaves=llaves, rep=rep):
            if len(muestra) < args.filas:
                muestra.append(fila)
            if llaves is not None:
                cubiertas.add((fila.get("sku"), fila.get("tienda"), fila.get("fecha")))

        print(f"  filas leídas    {rep.filas_leidas:,}")
        print(f"  filas devueltas {rep.filas_devueltas:,}")
        if rep.totales_descartados:
            print(f"  filas de total descartadas {rep.totales_descartados}")
        if rep.sin_fecha:
            print(f"  filas sin fecha legible    {rep.sin_fecha:,}")
        print(f"  rango de fechas {rep.desde} -> {rep.hasta}")
        print(f"  tiendas         {sorted(rep.tiendas)[:10]}")
        for a in rep.advertencias:
            print(f"  ! {a}")
        for m in muestra:
            print(f"    {m}")

        if llaves:
            pct = len(cubiertas) / len(llaves) * 100
            print(f"  CRUCE: cubre {len(cubiertas):,} de {len(llaves):,} días "
                  f"con faltante = {pct:.1f}%")
            if pct == 0:
                print("  ! Cero cruce. Revisar que sea la misma tienda y el mismo "
                      "periodo que BOPS_OSA antes de seguir.")

    return 0


if __name__ == "__main__":
    sys.exit(main())
