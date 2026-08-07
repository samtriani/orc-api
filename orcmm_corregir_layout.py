"""
ORCMM - Corrector de layout.

Toma un Excel de captura ya llenado por los equipos y lo deja alineado con la
especificación de orcmm_layout_spec.py, SIN tocar los datos de negocio.

    python orcmm_corregir_layout.py "040826_La Comer_Layout de datos RCA (OSA)_V2_Con Datos.xlsx"

Qué corrige:

  - Encabezados con nombre de negocio en vez de nombre de campo
    ("# de Cita" -> folio_cita). El pipeline ya los traduce al vuelo; esto los
    alinea en el origen para que el archivo y el modelo digan lo mismo.
  - Filas guía 1, 2, 4 y 5 de cada hoja, regeneradas desde el spec: título,
    fuente/owner/ventana, tipo esperado con su asterisco y descripción.
  - Columnas declaradas en el spec que no vienen en el archivo: se agregan
    vacías al final, con su guía, para que los equipos las puedan llenar.
  - Claves capturadas como número (tienda, cedis, folio) pasadas a TEXTO, que
    es lo que evita perder ceros a la izquierda y la notación científica.
  - Fórmulas en las celdas de datos, congeladas a su valor calculado. Un XLOOKUP
    que jala de otra hoja se ve bien en Excel, pero al guardar el archivo desde
    cualquier herramienta que no sea Excel el valor en caché se pierde y la
    columna llega vacía al modelo. Además el layout pide el registro tal como
    sale del sistema, no una derivación dentro del propio archivo.
  - La hoja INSTRUCCIONES: reglas de llenado y tabla de responsables.

Qué NO corrige, porque no se puede inventar:

  - Filas faltantes (los 23 pedidos sin cita, la hoja SIMA vacía).
  - Datos que se contradicen entre hojas.
  - Columnas constantes que no distinguen nada (estatus_cita).

Esas salen en orcmm_validar_layout.py y se piden al equipo dueño de la hoja.

El archivo original NO se modifica: se escribe una copia corregida.
"""

from __future__ import annotations

import argparse
import sys
from copy import copy
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from orcmm_layout_spec import (CSV, FILA_DATOS, FILA_DESCRIPCION, FILA_ENCABEZADO,
                               FILA_META, FILA_TIPO, FILA_TITULO, HOJAS,
                               REGLAS_DE_LLENADO, normalizar_encabezado, origen_de)
from orcmm_validar_layout import CAMPOS_TEXTO_CLAVE

# Si ninguna hoja del archivo trae las filas 1 y 2 bien puestas, se usan estos.
FUENTE_TITULO = Font(bold=True, size=14)
FUENTE_META = Font(size=9, color="595959")


def _clave_a_texto(v) -> Optional[str]:
    """287.0 -> '287'. Deja intacto lo que ya es texto."""
    if v is None or isinstance(v, str):
        return None
    if isinstance(v, bool) or not isinstance(v, (int, float)):
        return None
    return str(int(v)) if float(v).is_integer() else str(v)


def _copiar_estilo(destino, origen) -> None:
    if origen is not None:
        destino._style = copy(origen._style)


def _hoja_de_referencia(wb):
    """Una hoja del archivo que ya tenga bien las filas 1 y 2, para copiarle
    el formato en vez de adivinarlo."""
    for nombre in HOJAS:
        if nombre in wb.sheetnames:
            ws = wb[nombre]
            if ws.cell(row=FILA_TITULO, column=1).value and \
                    ws.cell(row=FILA_META, column=1).value:
                return ws
    return None


def _texto_guia(campo) -> tuple[str, str]:
    """Filas 4 y 5 de una columna: tipo esperado y descripción con ejemplo."""
    nombre, tipo, obligatorio, descripcion, ejemplo = campo
    fila4 = f"{tipo} *" if obligatorio else tipo
    fila5 = " ".join(x for x in (descripcion, f"Ej.: {ejemplo}" if ejemplo else "") if x)
    return fila4, fila5


def congelar_formulas(ws, ws_valores, nombre: str, cambios: List[str]) -> None:
    """Sustituye cada fórmula de las filas de datos por su resultado.

    openpyxl no reescribe el valor en caché de una fórmula: si el archivo se
    guarda fuera de Excel, la celda queda con la fórmula y sin resultado, y el
    pipeline la lee como SIN DATO. Congelar el valor deja el dato a salvo de
    quien vuelva a tocar el archivo.
    """
    congeladas, sin_valor = 0, 0
    for fila in ws.iter_rows(min_row=FILA_DATOS):
        for celda in fila:
            if not (isinstance(celda.value, str) and celda.value.startswith("=")):
                continue
            valor = ws_valores[celda.coordinate].value
            if valor is None:
                sin_valor += 1     # nunca se recalculó: se conserva la fórmula
                continue
            celda.value = valor
            congeladas += 1

    if congeladas:
        cambios.append(f"{nombre}: {congeladas} fórmulas congeladas a su valor calculado")
    if sin_valor:
        cambios.append(f"{nombre}: {sin_valor} fórmulas SIN resultado en caché, se dejaron "
                       f"tal cual. Abrir el archivo en Excel, guardarlo y volver a correr "
                       f"esta corrección, o el modelo leerá esas celdas vacías.")


def corregir_hoja(wb, nombre: str, referencia, cambios: List[str], ws_valores=None) -> None:
    ws = wb[nombre]
    spec = HOJAS[nombre]
    campo_de = {c[0]: c for c in spec["campos"]}

    if ws_valores is not None:
        congelar_formulas(ws, ws_valores, nombre, cambios)

    # --- encabezados: nombre de negocio -> nombre de campo ----------------
    encabezados: List[Optional[str]] = []
    for celda in ws[FILA_ENCABEZADO]:
        crudo = celda.value
        canonico = normalizar_encabezado(nombre, crudo)
        if canonico and crudo != canonico:
            celda.value = canonico
            cambios.append(f"{nombre}: encabezado '{crudo}' -> '{canonico}'")
        encabezados.append(canonico)

    # --- columnas del spec que no vienen en el archivo --------------------
    ausentes = [c for c in spec["campos"] if c[0] not in encabezados]
    if ausentes:
        col = len(encabezados) + 1
        for campo in ausentes:
            for fila, valor in ((FILA_ENCABEZADO, campo[0]),
                                (FILA_TIPO, _texto_guia(campo)[0]),
                                (FILA_DESCRIPCION, _texto_guia(campo)[1])):
                celda = ws.cell(row=fila, column=col, value=valor)
                _copiar_estilo(celda, ws.cell(row=fila, column=col - 1))
            ws.column_dimensions[get_column_letter(col)].width = \
                ws.column_dimensions[get_column_letter(col - 1)].width or 18
            encabezados.append(campo[0])
            cambios.append(f"{nombre}: columna nueva '{campo[0]}' "
                           f"(opcional, vacía) en {get_column_letter(col)}")
            col += 1

    # --- filas guía 1 y 2 -------------------------------------------------
    # El separador es el punto medio que ya usa el archivo, no un guión: así
    # las hojas que estaban bien no se reescriben nada más por el formato.
    titulo = f"{spec['equipo'].upper()} · {spec['titulo']}"
    meta = (f"{spec['forma']}    |    Owner: {spec['owner']}    |    "
            f"Ventana: {spec['ventana']}")
    for fila, valor, fuente in ((FILA_TITULO, titulo, FUENTE_TITULO),
                                (FILA_META, meta, FUENTE_META)):
        celda = ws.cell(row=fila, column=1)
        if celda.value == valor:
            continue
        vacia = celda.value is None
        celda.value = valor
        if referencia is not None and referencia is not ws:
            _copiar_estilo(celda, referencia.cell(row=fila, column=1))
        else:
            celda.font = fuente
        cambios.append(f"{nombre}: fila {fila} {'escrita' if vacia else 'actualizada'} "
                       f"({'título' if fila == FILA_TITULO else 'fuente / owner / ventana'})")

    # --- filas guía 4 y 5, por columna ------------------------------------
    for idx, canonico in enumerate(encabezados, 1):
        campo = campo_de.get(canonico)
        if campo is None:
            continue
        for fila, valor in zip((FILA_TIPO, FILA_DESCRIPCION), _texto_guia(campo)):
            celda = ws.cell(row=fila, column=idx)
            if celda.value != valor:
                anterior = celda.value
                celda.value = valor
                if fila == FILA_TIPO:
                    cambios.append(f"{nombre}.{canonico}: tipo '{anterior}' -> '{valor}'")

    # --- claves numéricas a texto -----------------------------------------
    convertidas: dict[str, int] = {}
    for idx, canonico in enumerate(encabezados, 1):
        if canonico not in CAMPOS_TEXTO_CLAVE:
            continue
        for fila in range(FILA_DATOS, ws.max_row + 1):
            celda = ws.cell(row=fila, column=idx)
            texto = _clave_a_texto(celda.value)
            if texto is None:
                continue
            celda.value = texto
            celda.number_format = "@"
            convertidas[canonico] = convertidas.get(canonico, 0) + 1
    for campo, n in convertidas.items():
        cambios.append(f"{nombre}.{campo}: {n} celdas de número a TEXTO")

    if ws.freeze_panes != f"A{FILA_DATOS}":
        ws.freeze_panes = f"A{FILA_DATOS}"
        cambios.append(f"{nombre}: paneles congelados en A{FILA_DATOS}")
    ws.sheet_view.showGridLines = False


def corregir_instrucciones(wb, cambios: List[str]) -> None:
    """Deja la portada consistente con las hojas que realmente existen."""
    if "INSTRUCCIONES" not in wb.sheetnames:
        return
    ws = wb["INSTRUCCIONES"]

    # "Estas ocho hojas..." se quedó corto al agregar la de citas.
    celda = ws["C7"]
    if isinstance(celda.value, str):
        for viejo in ("ocho hojas", "8 hojas"):
            if viejo in celda.value:
                celda.value = celda.value.replace(viejo, f"{len(HOJAS)} hojas")
                cambios.append(f"INSTRUCCIONES: '{viejo}' -> '{len(HOJAS)} hojas'")

    # Reglas de llenado: fila 12 en adelante, una por regla del spec.
    fila = 12
    for titulo, detalle in REGLAS_DE_LLENADO:
        if ws.cell(row=fila, column=2).value != titulo:
            _copiar_estilo(ws.cell(row=fila, column=2), ws.cell(row=fila - 1, column=2))
            _copiar_estilo(ws.cell(row=fila, column=3), ws.cell(row=fila - 1, column=3))
            cambios.append(f"INSTRUCCIONES: regla de llenado '{titulo}' en la fila {fila}")
        ws.cell(row=fila, column=2, value=titulo)
        ws.cell(row=fila, column=3, value=detalle)
        fila += 1

    # Tabla de responsables: una fila por hoja del spec, desde la 23.
    fila = 23
    for nombre, spec in HOJAS.items():
        if ws.cell(row=fila, column=2).value != nombre:
            for col in range(2, 6):
                _copiar_estilo(ws.cell(row=fila, column=col),
                               ws.cell(row=fila - 1, column=col))
            cambios.append(f"INSTRUCCIONES: hoja '{nombre}' agregada a la tabla de equipos")
        ws.cell(row=fila, column=2, value=nombre)
        ws.cell(row=fila, column=3, value=spec["para_que"])
        ws.cell(row=fila, column=4, value=spec["equipo"])
        ws.cell(row=fila, column=5, value=spec["owner"])
        fila += 1


def corregir(entrada: Path, salida: Path) -> List[str]:
    wb = openpyxl.load_workbook(entrada)
    # La misma hoja leída dos veces: una con las fórmulas y otra con lo que
    # esas fórmulas dieron la última vez que Excel las calculó.
    wb_valores = openpyxl.load_workbook(entrada, data_only=True)
    cambios: List[str] = []
    referencia = _hoja_de_referencia(wb)

    for nombre in HOJAS:
        if nombre not in wb.sheetnames:
            if origen_de(nombre) == CSV:
                continue      # se entrega como archivo aparte, no como hoja
            cambios.append(f"{nombre}: la hoja no existe en el archivo, no se puede corregir "
                           f"desde aquí. Hay que agregarla.")
            continue
        corregir_hoja(wb, nombre, referencia, cambios, wb_valores[nombre])

    corregir_instrucciones(wb, cambios)
    wb.save(salida)
    return cambios


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Alinea un Excel de captura con el layout del spec, sin tocar los datos.")
    ap.add_argument("archivo")
    ap.add_argument("-o", "--salida", default=None,
                    help="Por omisión, '<archivo> corregido.xlsx' junto al original.")
    args = ap.parse_args()

    entrada = Path(args.archivo)
    if not entrada.exists():
        print(f"No se encontró el archivo: {entrada}", file=sys.stderr)
        return 1

    salida = Path(args.salida) if args.salida else \
        entrada.with_name(f"{entrada.stem} corregido.xlsx")
    if salida.resolve() == entrada.resolve():
        print("La salida no puede ser el archivo de entrada: el original se conserva.",
              file=sys.stderr)
        return 1

    cambios = corregir(entrada, salida)

    print(f"\nCorrigiendo  {entrada.name}")
    if cambios:
        print(f"\n{len(cambios)} cambios de layout:")
        for c in cambios:
            print(f"  - {c}")
    else:
        print("\nEl layout ya estaba alineado con el spec.")

    print(f"\nOriginal intacto:  {entrada.name}")
    print(f"Archivo corregido: {salida.name}")
    print("\nSiguiente paso:    python orcmm_validar_layout.py "
          f"\"{salida.name}\"\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
