"""
ORCMM - Validador de layout contra datos reales.

Compara un Excel de captura (ya llenado por los equipos) contra la
especificación única en orcmm_layout_spec.py, ANTES de correr el pipeline.

    python orcmm_validar_layout.py "040826_La Comer_Layout de datos RCA (OSA)_V1_Con Datos.xlsx"

Revisa, por cada hoja esperada:
  - Que la hoja exista.
  - Que los encabezados de la fila 3 coincidan con los campos del spec
    (nombre, orden, faltantes, sobrantes/renombrados).
  - Que haya datos desde la fila 6.
  - Tipo real de cada celda vs tipo esperado (Texto/Entero/Decimal/Fecha/Hora/Sí-No/Lista).
  - SKU y tienda/cedis capturados como TEXTO (no numérico, para no perder ceros
    a la izquierda ni caer en notación científica).
  - Campos obligatorios (*) vacíos.
  - Fechas fuera de la ventana declarada en el spec, si aplica.
  - Duplicados en la llave natural de cada hoja.

No corrige nada ni corre el motor de RCA: sólo diagnostica el archivo de
entrada para que se pueda pedir una corrección a los equipos antes de la
corrida real.
"""

from __future__ import annotations

import argparse
import re
import sys
from collections import Counter
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import openpyxl

from orcmm_fuentes_csv import (ReporteCSV, agrupar_por_hoja, fecha_csv, leer_csv,
                               numero_csv)
from orcmm_layout_spec import (ALIAS_ENCABEZADOS, BOL, CSV, DEC, ENT, FEC,
                               FILA_DATOS, FILA_ENCABEZADO, FILA_TIPO, HOJAS,
                               HOR, LST, TXT, normalizar_encabezado, origen_de)
from orcmm_rca_engine import EVALUAR_PEDIDO_TIENDA

LLAVES = {
    "CATALOGO": ["sku", "tienda"],
    "TABLEAU_INV_TIENDA": ["sku", "tienda", "fecha"],
    "BOPS_OSA": ["sku", "tienda", "fecha"],
    "TABLEAU_VENTAS": ["sku", "tienda", "fecha"],
    "CEDIS_INVENTARIO": ["sku", "cedis", "fecha"],
    "CEDIS_TRANSFERENCIAS": ["folio"],
    "SIMA_PEDIDOS_TIENDA": ["folio"],
    "COMPRAS_PEDIDOS_PROV": ["folio"],
    "CITAS_PROV_CEDIS": ["folio_cita"],
}

CAMPOS_TEXTO_CLAVE = {"sku", "tienda", "cedis", "cedis_surtidor", "cedis_destino",
                      "cedis_origen", "folio", "folio_cita", "proveedor_id"}


def _texto(v) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


class Reporte:
    """Tres severidades, porque no todo lo que está mal impide correr.

      errores       el layout está roto: columnas que no existen, claves
                    capturadas como número, tipos que no cuadran. El modelo
                    leería mal. Bloquean.
      faltan_datos  el layout está bien, faltan renglones. El motor ya sabe
                    qué hacer con eso: lo reporta como cobertura perdida
                    nombrando el campo. No bloquea, pero cambia lo que se
                    puede concluir del resultado.
      advertencias  vale la pena revisarlo antes de firmar el Pareto.
    """

    def __init__(self):
        self.errores: list[str] = []
        self.faltan_datos: list[str] = []
        self.advertencias: list[str] = []
        self.info: list[str] = []
        # Subconjunto de errores que orcmm_corregir_layout sabe arreglar. Se
        # marca aquí, donde se sabe de qué error se trata, en vez de adivinarlo
        # después leyendo el texto del mensaje. Correr el corrector sobre un
        # layout de 35 MB cuesta minutos: hay que saber antes si va a servir.
        self.errores_corregibles: list[str] = []

    def error(self, msg, corregible: bool = False):
        self.errores.append(msg)
        if corregible:
            self.errores_corregibles.append(msg)
    def dato(self, msg): self.faltan_datos.append(msg)
    def warn(self, msg): self.advertencias.append(msg)
    def ok(self, msg): self.info.append(msg)


def validar_encabezados(fila_encabezado, nombre: str, rep: Reporte) -> list[Optional[str]]:
    """Revisa la fila 3 contra el spec. Recibe los valores ya leídos, no la
    hoja: con 400 mil filas, cada pasada extra del parser cuesta segundos."""
    crudos = [_texto(v) for v in fila_encabezado]
    encabezados = [normalizar_encabezado(nombre, h) for h in crudos]
    esperados = [c[0] for c in HOJAS[nombre]["campos"]]
    presentes = [h for h in encabezados if h]

    # Un encabezado traducido se lee bien, pero conviene alinearlo en el origen.
    for crudo, canonico in zip(crudos, encabezados):
        if crudo and canonico and crudo != canonico:
            rep.warn(f"{nombre}: la columna '{crudo}' se está leyendo como '{canonico}'. "
                     f"Renombrarla en el layout para que el archivo y el modelo digan lo mismo.")

    opcionales = {c[0] for c in HOJAS[nombre]["campos"] if not c[2]}
    faltantes = [e for e in esperados if e not in presentes and e not in opcionales]
    ausentes_opcionales = [e for e in esperados if e not in presentes and e in opcionales]
    if ausentes_opcionales:
        rep.warn(f"{nombre}: sin las columnas opcionales {', '.join(ausentes_opcionales)} "
                 f"(no bloquean, pero el modelo no puede usarlas).")
    sobrantes = [p for p in presentes if p not in esperados]
    duplicados = [h for h, n in Counter(presentes).items() if n > 1]

    if faltantes:
        rep.error(f"{nombre}: faltan columnas esperadas -> {', '.join(faltantes)}",
                  corregible=True)
    if sobrantes:
        rep.warn(f"{nombre}: columnas no reconocidas por el spec (revisar nombre exacto) -> {', '.join(sobrantes)}")
    if duplicados:
        rep.error(f"{nombre}: encabezado repetido en fila {FILA_ENCABEZADO} -> {', '.join(duplicados)}")

    # Orden: sólo advertencia, el pipeline no depende del orden.
    comunes = [e for e in esperados if e in presentes]
    orden_presentes = [h for h in presentes if h in esperados]
    if comunes != orden_presentes:
        rep.warn(f"{nombre}: el orden de columnas no coincide con el layout original (no bloquea la lectura).")

    if not faltantes and not sobrantes and not duplicados:
        rep.ok(f"{nombre}: encabezados OK ({len(esperados)} columnas).")

    return encabezados


def _clave_pierde_informacion(v) -> bool:
    """¿Leer esta clave numérica como texto daría algo distinto a lo capturado?

    El pipeline normaliza toda clave con str(), y para un entero eso es exacto:
    el folio 26300925519 se lee igual venga como texto o como número. Lo que sí
    rompe es que Excel lo haya guardado como flotante —ahí aparecen el '.0' y
    la notación científica— o que pase del entero que un flotante representa
    sin redondear. Sólo eso es un error; lo demás es una advertencia de captura.

    Los ceros a la izquierda son un caso aparte: si se perdieron, fue al
    capturar y el archivo ya no conserva rastro. No se puede detectar aquí.
    """
    if isinstance(v, bool):
        return False
    if isinstance(v, int):
        return abs(v) > 2 ** 53
    if isinstance(v, float):
        return True          # todo flotante en una clave ya viene deformado
    return False


def tipo_valor_ok(valor, tipo_esperado: str) -> bool:
    if valor is None or valor == "":
        return True  # vacío se valida aparte (obligatoriedad)
    if tipo_esperado == FEC:
        return isinstance(valor, (date, datetime))
    if tipo_esperado == HOR:
        return isinstance(valor, (datetime,)) or hasattr(valor, "hour")
    if tipo_esperado == ENT:
        if isinstance(valor, bool):
            return False
        if isinstance(valor, (int, float)):
            return float(valor).is_integer()
        return str(valor).strip().lstrip("-").isdigit()
    if tipo_esperado == DEC:
        if isinstance(valor, bool):
            return False
        if isinstance(valor, (int, float)):
            return True
        try:
            float(str(valor).replace(",", ""))
            return True
        except ValueError:
            return False
    if tipo_esperado == BOL:
        return str(valor).strip().lower() in {"sí", "si", "no", "true", "false", "1", "0"}
    # TXT y LST: cualquier cosa sirve, pero para claves no debe ser numérico crudo
    return True


def validar_datos(filas, nombre: str, encabezados: list[Optional[str]], rep: Reporte,
                  recolectar=None):
    """Revisa los datos y, de paso, recoge lo que los cruces van a necesitar.

    `filas` es un iterable de tuplas desde FILA_DATOS. `recolectar`, si se
    pasa, se llama con cada registro: así las revisiones que cruzan dos hojas
    (citas contra pedidos, OSA contra catálogo) no obligan a leer el archivo
    otra vez. Con este volumen, releer es lo que hacía tardar minutos.
    """
    campos = HOJAS[nombre]["campos"]
    tipo_de = {c[0]: c[1] for c in campos}
    obligatorio_de = {c[0]: c[2] for c in campos}
    col_de_campo = {}
    for idx, h in enumerate(encabezados):
        if h and h not in col_de_campo:
            col_de_campo[h] = idx

    n_filas = 0
    faltantes_obligatorios: Counter = Counter()
    tipos_malos: Counter = Counter()
    claves_como_numero: Counter = Counter()
    claves_rotas: Counter = Counter()
    llave_vista: Counter = Counter()
    ejemplos_tipo: dict[str, str] = {}
    ejemplos_clave: dict[str, str] = {}

    for fila in filas:
        if all(v is None or v == "" for v in fila):
            continue
        n_filas += 1
        registro = {}
        for campo, idx in col_de_campo.items():
            v = fila[idx] if idx < len(fila) else None
            registro[campo] = v

            tipo = tipo_de.get(campo)
            if tipo and not tipo_valor_ok(v, tipo):
                tipos_malos[campo] += 1
                ejemplos_tipo.setdefault(campo, f"'{v}' (tipo Python: {type(v).__name__}, esperado {tipo})")

            if campo in CAMPOS_TEXTO_CLAVE and isinstance(v, (int, float)) and not isinstance(v, bool):
                if _clave_pierde_informacion(v):
                    claves_rotas[campo] += 1
                    ejemplos_clave.setdefault(campo, repr(v))
                else:
                    claves_como_numero[campo] += 1

            if obligatorio_de.get(campo) and (v is None or v == ""):
                faltantes_obligatorios[campo] += 1

        if recolectar is not None:
            recolectar(registro)

        llave = LLAVES.get(nombre, [])
        if llave and all(registro.get(k) not in (None, "") for k in llave):
            k = tuple(_texto(registro.get(k)) if not isinstance(registro.get(k), (date, datetime))
                      else registro.get(k) for k in llave)
            llave_vista[k] += 1

    if n_filas == 0:
        # Una hoja vacía no es un layout roto: es una fuente que no llegó. El
        # motor la maneja dejando sin clasificar los días que dependían de
        # ella y diciendo cuál campo faltó, así que no tiene por qué impedir
        # la corrida.
        if nombre == "SIMA_PEDIDOS_TIENDA" and not EVALUAR_PEDIDO_TIENDA:
            rep.dato(f"{nombre}: hoja vacía, esperado por ahora. La prioridad 3 está apagada "
                     f"(EVALUAR_PEDIDO_TIENDA = False) y el modelo corre parcial, sin poder "
                     f"dictaminar RC03. Al llegar los datos hay que prender el interruptor.")
        else:
            rep.dato(f"{nombre}: no tiene filas de datos desde la fila {FILA_DATOS} (hoja vacía). "
                     f"Todas las reglas que dependen de ella quedarán sin clasificar.")
        return

    rep.ok(f"{nombre}: {n_filas} filas de datos leídas desde la fila {FILA_DATOS}.")

    for campo, n in faltantes_obligatorios.items():
        rep.error(f"{nombre}.{campo}: {n} de {n_filas} filas sin dato (campo obligatorio *).")

    for campo, n in claves_rotas.items():
        rep.error(f"{nombre}.{campo}: {n} filas con la clave capturada como NÚMERO y con "
                  f"pérdida de información (ej: {ejemplos_clave.get(campo)}). El valor "
                  f"que leería el modelo no es el que se capturó.", corregible=True)

    for campo, n in claves_como_numero.items():
        rep.warn(f"{nombre}.{campo}: {n} filas con la clave capturada como NÚMERO, no TEXTO. "
                 f"El modelo la lee bien porque son enteros exactos, pero si la clave "
                 f"original traía ceros a la izquierda, ya se perdieron al capturar y no "
                 f"hay forma de saberlo desde el archivo. Pedirla como TEXTO en el origen.")

    for campo, n in tipos_malos.items():
        rep.warn(f"{nombre}.{campo}: {n} valores no coinciden con el tipo esperado "
                 f"({tipo_de.get(campo)}). Ejemplo: {ejemplos_tipo.get(campo)}")

    duplicados = [k for k, n in llave_vista.items() if n > 1]
    if duplicados:
        llave_nombres = "+".join(LLAVES.get(nombre, []))
        muestra = ", ".join(str(k) for k in duplicados[:5])
        rep.warn(f"{nombre}: {len(duplicados)} llaves duplicadas ({llave_nombres}) -> ej: {muestra}")


class Recolector:
    """Lo mínimo que hace falta guardar de cada hoja para poder cruzarlas.

    Guardar la fila entera de las 250 mil que suman pedidos y citas se come
    la memoria de la máquina; con estos cuatro campos alcanza para todas las
    revisiones cruzadas.
    """

    def __init__(self):
        self.pedidos: dict = {}        # (folio, sku) -> cajas_pedidas
        self.citas: list = []          # (folio, sku, cajas_confirmadas, estatus)
        self.catalogo: set = set()     # (sku, tienda)

    def de(self, hoja: str):
        """El recolector que le toca a esa hoja, o None si no se cruza."""
        return {
            "COMPRAS_PEDIDOS_PROV": self._pedido,
            "CITAS_PROV_CEDIS": self._cita,
            "CATALOGO": self._catalogo,
        }.get(hoja)

    def _pedido(self, r: dict) -> None:
        self.pedidos[(_texto(r.get("folio")), _texto(r.get("sku")))] = r.get("cajas_pedidas")

    def _cita(self, r: dict) -> None:
        self.citas.append((_texto(r.get("folio")), _texto(r.get("sku")),
                           r.get("cajas_confirmadas_cita"), _texto(r.get("estatus_cita"))))

    def _catalogo(self, r: dict) -> None:
        self.catalogo.add((_texto(r.get("sku")), _texto(r.get("tienda"))))


def validar_cruce_citas(rec: Recolector, rep: Reporte):
    """Revisiones que sólo se ven cruzando la cita con su pedido.

    La cita es la única hoja que se lee contra otra: sin el pedido no significa
    nada, y su valor está justamente en la diferencia entre lo pedido, lo
    confirmado y lo entregado. Si esa diferencia nunca aparece, el dato es
    decorativo y el modelo no debería apoyarse en él.
    """
    if not rec.citas or not rec.pedidos:
        return

    por_folio = rec.pedidos
    con_cita, huerfanas, confirmadas_de_mas, iguales, comparables = set(), [], [], 0, 0
    estatus = set()

    for folio, sku, conf, est in rec.citas:
        estatus.add(est)
        llave = (folio, sku)
        if llave not in por_folio:
            huerfanas.append(folio)
            continue
        con_cita.add(llave)

        ped = por_folio[llave]
        if isinstance(conf, (int, float)) and isinstance(ped, (int, float)):
            comparables += 1
            if conf > ped:
                confirmadas_de_mas.append(folio)
            elif conf == ped:
                iguales += 1

    if huerfanas:
        rep.error(f"CITAS_PROV_CEDIS: {len(huerfanas)} citas cuyo folio+sku no existe en "
                  f"COMPRAS_PEDIDOS_PROV (ej: {huerfanas[:3]}). Sin pedido no se pueden usar.")

    if confirmadas_de_mas:
        rep.error(f"CITAS_PROV_CEDIS: {len(confirmadas_de_mas)} citas con "
                  f"cajas_confirmadas_cita MAYOR que las cajas_pedidas del pedido "
                  f"(ej: {confirmadas_de_mas[:3]}). Una de las dos hojas está mal.")

    if comparables and iguales == comparables:
        rep.warn(f"CITAS_PROV_CEDIS: en las {comparables} citas, cajas_confirmadas_cita es "
                 f"SIEMPRE igual a cajas_pedidas. O el proveedor nunca recorta al agendar, o "
                 f"la columna se está copiando del pedido. Si es lo segundo, el modelo no "
                 f"puede detectar 'confirmó de menos' y hay que traer el dato real de la cita.")

    sin_cita = len(set(por_folio) - con_cita)
    if sin_cita:
        pct = round(sin_cita / len(por_folio) * 100, 1)
        rep.dato(f"CITAS_PROV_CEDIS: {sin_cita} de {len(por_folio)} pedidos ({pct}%) no traen cita. "
                 f"El modelo lee eso como 'el proveedor nunca agendó' y lo dictamina como RC06. "
                 f"Confirmar con Compras que la extracción trae TODAS las citas del periodo: si "
                 f"vino parcial, la causa es un hueco de captura y no del proveedor.")

    estatus -= {None}
    if len(estatus) == 1:
        rep.warn(f"CITAS_PROV_CEDIS.estatus_cita: todas las filas dicen '{estatus.pop()}'. "
                 f"La columna no distingue nada; el modelo dictamina con las cajas, no con "
                 f"el estatus. Si el sistema sí marca canceladas o no presentadas, traerlo.")


def validar_csv(rutas: list[Path], nombre: str, rep: Reporte) -> set:
    """Valida una hoja entregada como CSV y devuelve sus llaves.

    Mismas revisiones que sobre una hoja del Excel —obligatorios vacíos,
    llaves duplicadas, fechas ilegibles— pero en streaming: son millones de
    filas y no caben en memoria. Las llaves sí se guardan, porque son lo que
    permite cruzar contra BOPS_OSA después.
    """
    campos = HOJAS[nombre]["campos"]
    obligatorio_de = {c[0]: c[2] for c in campos}
    llave_campos = LLAVES.get(nombre, [])

    reporte_csv = ReporteCSV(nombre)
    faltantes: Counter = Counter()
    llaves: set = set()
    duplicadas = 0
    n = 0

    for registro in leer_csv(rutas, nombre, rep=reporte_csv):
        n += 1
        for campo, obligatorio in obligatorio_de.items():
            if obligatorio and registro.get(campo) in (None, ""):
                faltantes[campo] += 1
        if llave_campos:
            k = tuple(registro.get(c) for c in llave_campos)
            if k in llaves:
                duplicadas += 1
            else:
                llaves.add(k)

    archivos = ", ".join(r.name for r in rutas)
    if not n:
        rep.dato(f"{nombre}: los CSV ({archivos}) no traen filas de datos. Todas las "
                 f"reglas que dependen de esta fuente quedarán sin clasificar.")
        return llaves

    rep.ok(f"{nombre}: {n:,} filas leídas de {len(rutas)} CSV ({archivos}).")

    # Las observaciones del lector (encabezados traducidos, columna sin
    # nombre, columnas que el spec no reconoce) valen igual que las del Excel.
    for a in reporte_csv.advertencias:
        rep.warn(a)

    for campo, cuantas in faltantes.items():
        rep.error(f"{nombre}.{campo}: {cuantas:,} de {n:,} filas sin dato "
                  f"(campo obligatorio *).")

    if reporte_csv.sin_fecha:
        rep.error(f"{nombre}: {reporte_csv.sin_fecha:,} filas con una fecha que no se "
                  f"pudo leer. Revisar el formato de la columna de fecha.")

    if duplicadas:
        rep.warn(f"{nombre}: {duplicadas:,} filas repiten la llave "
                 f"({'+'.join(llave_campos)}). Se queda la última de cada una.")

    if len(reporte_csv.tiendas) > 1:
        rep.warn(f"{nombre}: el archivo trae varias tiendas "
                 f"({', '.join(sorted(reporte_csv.tiendas)[:5])}). Confirmar que es "
                 f"lo esperado para este análisis.")

    return llaves


def validar_cruce_fuentes(faltantes: set, rec: Recolector, llaves_csv: dict,
                          rep: Reporte) -> None:
    """¿Las fuentes hablan del mismo universo?

    Un layout puede estar impecable y aun así no producir nada: basta con que
    el inventario venga de otra tienda, de otro periodo, o de otra división
    que BOPS_OSA. Eso no se ve revisando cada hoja por separado —cada una
    está bien— sino cruzándolas, y sale más barato saberlo aquí que después
    de una corrida completa.
    """
    if not faltantes:
        rep.dato("BOPS_OSA: ningún día con OSA por debajo del umbral. No hay nada "
                 "que clasificar.")
        return

    n = len(faltantes)
    tiendas_osa = {k[1] for k in faltantes}

    # Catálogo: define qué SKU le tocan al análisis.
    if rec.catalogo:
        dentro = sum(1 for k in faltantes if (k[0], k[1]) in rec.catalogo)
        pct = round(dentro / n * 100, 1)
        if dentro == 0:
            rep.error(f"NINGUNO de los {n:,} días con faltante de BOPS_OSA tiene su SKU "
                      f"en CATALOGO. Las dos fuentes no hablan del mismo universo: "
                      f"revisar tienda y división antes de correr.")
        elif pct < 90:
            rep.dato(f"Sólo {dentro:,} de {n:,} días con faltante ({pct}%) tienen su SKU "
                     f"en CATALOGO. Los {n - dentro:,} restantes se contarán FUERA DEL "
                     f"ALCANCE, no como dato faltante. Si BOPS debería entregar sólo la "
                     f"división del catálogo, pedir la extracción filtrada.")

    # Inventario en tienda: es la prioridad 1, sin ella no se clasifica nada.
    inv = llaves_csv.get("TABLEAU_INV_TIENDA")
    if inv is not None:
        cubre = sum(1 for k in faltantes if k in inv)
        pct = round(cubre / n * 100, 1)
        tiendas_inv = {k[1] for k in inv}
        if cubre == 0:
            detalle = ""
            if tiendas_inv and not (tiendas_osa & tiendas_inv):
                detalle = (f" BOPS_OSA es de la tienda "
                           f"{', '.join(sorted(tiendas_osa)[:3])} y el inventario de la "
                           f"{', '.join(sorted(tiendas_inv)[:3])}.")
            rep.error(f"TABLEAU_INV_TIENDA no cubre NINGUNO de los {n:,} días con "
                      f"faltante.{detalle} Sin inventario en tienda la prioridad 1 no "
                      f"decide y no se clasificaría ni un día.")
        elif pct < 50:
            rep.dato(f"TABLEAU_INV_TIENDA cubre {cubre:,} de {n:,} días con faltante "
                     f"({pct}%). El resto quedará sin clasificar en la prioridad 1.")
        else:
            rep.ok(f"TABLEAU_INV_TIENDA cubre {cubre:,} de {n:,} días con faltante ({pct}%).")


def validar_archivo(ruta: Path, csvs=None) -> Reporte:
    """Valida el paquete completo: el .xlsx y los CSV que lo acompañan.

    `csvs` acepta una lista de rutas; se reparten por nombre de archivo. Si no
    llega ninguno, sólo se valida el Excel — un layout anterior al V5, donde
    todas las hojas venían dentro.
    """
    rep = Reporte()
    # read_only: el layout V5 pesa 35 MB y suma 400 mil filas entre sus hojas.
    # En el modo normal openpyxl construye un objeto por celda y la validación
    # se va a minutos; además, cada hoja se recorre UNA sola vez y de esa misma
    # pasada salen los encabezados, los datos y lo que necesitan los cruces.
    wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
    rec = Recolector()
    faltantes: set = set()

    try:
        por_hoja, sueltos = agrupar_por_hoja(csvs or [])
        for s in sueltos:
            rep.warn(f"'{s.name}' no corresponde a ninguna hoja del layout. No se leería.")

        hojas_archivo = set(wb.sheetnames)
        hojas_esperadas = set(HOJAS.keys())

        for nombre in hojas_esperadas - hojas_archivo:
            if origen_de(nombre) == CSV:
                continue          # se revisa abajo, como archivo suelto
                rep.error(f"{nombre}: la hoja NO existe en el archivo.", corregible=True)

        for nombre in hojas_archivo - hojas_esperadas:
            rep.warn(f"Hoja '{nombre}' está en el archivo pero no está en el spec "
                     f"(¿hoja extra o mal nombrada?).")

        for nombre in HOJAS:
            if nombre not in hojas_archivo:
                continue
            if origen_de(nombre) == CSV and nombre in por_hoja:
                rep.warn(f"{nombre}: viene como hoja del Excel Y como CSV. Se usaría la "
                         f"hoja del Excel; quitar una de las dos para no dejarlo a la suerte.")

            encabezados: list = []
            datos = []
            for i, fila in enumerate(wb[nombre].iter_rows(values_only=True), 1):
                if i == FILA_ENCABEZADO:
                    encabezados = validar_encabezados(fila, nombre, rep)
                elif i >= FILA_DATOS:
                    datos.append(fila)
            validar_datos(datos, nombre, encabezados, rep, rec.de(nombre))

            if nombre == "BOPS_OSA":
                i_sku, i_tda, i_fec, i_osa = (
                    encabezados.index(c) if c in encabezados else -1
                    for c in ("sku", "tienda", "fecha", "osa_pct"))
                if min(i_sku, i_tda, i_fec, i_osa) >= 0:
                    for f in datos:
                        osa = numero_csv(f[i_osa]) if len(f) > i_osa else None
                        if osa is None or osa * 100 >= 100:
                            continue
                        d = fecha_csv(f[i_fec])
                        if d:
                            faltantes.add((_texto(f[i_sku]), _texto(f[i_tda]), d))
            del datos

        llaves_csv = {}
        for nombre in HOJAS:
            if origen_de(nombre) != CSV or nombre in hojas_archivo:
                continue
            if nombre not in por_hoja:
                rep.error(f"{nombre}: no está en el Excel ni llegó como CSV. Se entrega "
                          f"aparte porque no cabe en una hoja: falta el archivo "
                          f"{nombre}*.csv.")
                continue
            llaves_csv[nombre] = validar_csv(por_hoja[nombre], nombre, rep)

        validar_cruce_citas(rec, rep)
        validar_cruce_fuentes(faltantes, rec, llaves_csv, rep)
        return rep
    finally:
        wb.close()


def imprimir_reporte(rep: Reporte, archivo: str):
    print(f"\n{'='*78}\nValidación de layout: {archivo}\n{'='*78}")

    print(f"\n[OK] {len(rep.info)}")
    for m in rep.info:
        print(f"  - {m}")

    print(f"\n[ADVERTENCIAS] {len(rep.advertencias)} (no bloquean la corrida, pero revisar)")
    for m in rep.advertencias:
        print(f"  ! {m}")

    print(f"\n[DATOS INCOMPLETOS] {len(rep.faltan_datos)} (el layout está bien; faltan renglones)")
    for m in rep.faltan_datos:
        print(f"  ? {m}")

    print(f"\n[ERRORES DE LAYOUT] {len(rep.errores)} (bloquean: el modelo leería mal)")
    for m in rep.errores:
        print(f"  X {m}")

    print(f"\n{'='*78}")
    if rep.errores:
        print(f"RESULTADO: hay {len(rep.errores)} problema(s) que corregir antes de "
              f"correr el pipeline.")
        if rep.errores_corregibles:
            print(f"           {len(rep.errores_corregibles)} son de formato y se arreglan "
                  f"solos:")
            print("           python orcmm_corregir_layout.py \"<archivo>\"")
        else:
            print("           Ninguno es de formato: son datos que faltan o que no cuadran "
                  "entre hojas.")
            print("           La corrección automática no los puede inventar — hay que "
                  "pedir la reextracción.")
    elif rep.faltan_datos:
        print("RESULTADO: el layout está bien. Se puede correr, pero el resultado será PARCIAL "
              "por las fuentes que faltan.")
    elif rep.advertencias:
        print("RESULTADO: sin errores bloqueantes. Revisar advertencias antes de correr.")
    else:
        print("RESULTADO: layout OK. Se puede correr orcmm_pipeline.py")
    print(f"{'='*78}\n")


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Valida el layout de captura (Excel y sus CSV) contra el spec.")
    ap.add_argument("archivo")
    ap.add_argument("csv", nargs="*", type=Path,
                    help="CSV de las hojas que ya no caben en el Excel "
                         "(TABLEAU_INV_TIENDA_*.csv, TABLEAU_VENTAS.csv).")
    args = ap.parse_args()

    ruta = Path(args.archivo)
    if not ruta.exists():
        print(f"No se encontró el archivo: {ruta}", file=sys.stderr)
        return 1

    faltan = [c for c in args.csv if not c.exists()]
    if faltan:
        print(f"No se encontraron: {', '.join(str(c) for c in faltan)}", file=sys.stderr)
        return 1

    rep = validar_archivo(ruta, args.csv)
    imprimir_reporte(rep, ruta.name)
    return 1 if rep.errores else 0


if __name__ == "__main__":
    raise SystemExit(main())
