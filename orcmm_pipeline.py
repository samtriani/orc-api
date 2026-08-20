"""
ORCMM - Pipeline: del Excel de captura al resultado de causa raíz.

    python orcmm_pipeline.py "040826_La Comer_Layout de datos RCA (OSA)_V2_Con Datos.xlsx"

Qué hace, en orden:

  1. LEE     las 9 hojas del Excel que llenaron los equipos
  2. DERIVA  las banderas de cada día (¿había tránsito vigente? ¿pedido abierto?
             ¿la cita del proveedor ya venció?) a partir de los eventos con sus fechas
  3. CLASIFICA cada día de faltante con la matriz RC01-RC06 (orcmm_rca_engine)
  4. ESCRIBE un Excel de resultados con la clasificación diaria, el Pareto, el
             scorecard del proveedor y el reporte de cobertura

Principio: si una fuente viene vacía o le falta el día, la bandera queda en
None y el día se marca Sin clasificar nombrando el campo que falta. El
pipeline nunca rellena huecos con ceros.
"""

from __future__ import annotations

import argparse
import sys
import textwrap
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Set, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from orcmm_fuentes_csv import (ReporteCSV, agrupar_por_hoja, leer_csv,
                               llaves_con_faltante_ws)
from orcmm_layout_spec import (FILA_DATOS, FILA_ENCABEZADO, HOJAS,
                               ORIGEN_CENTRALIZADO, normalizar_encabezado)
from orcmm_rca_engine import (EVALUAR_PEDIDO_TIENDA, EXCLUIR_SKU_SIN_SIMA,
                              FUERA_DE_CATALOGO, SIN_DATO_SIMA,
                              EvidenciaSKUTienda, TipoResurtido, ViaResurtido)
from orcmm_rca_periodo import (clasificar, cobertura_modelo, dentro_del_alcance,
                               diagnosticar_periodo, pareto_periodo,
                               resumen_por_causa, resumen_por_responsable,
                               resumen_por_subcausa)

AMARILLO, GRIS, AZUL, AMBAR, VERDE, ROJO = "FFE600", "F2F2F2", "DDEBF7", "FFEB9C", "C6EFCE", "FFC7CE"
# Fuera de alcance: lila pálido. No puede ser otro gris — junto al de
# RC99 no se distinguirían, que es justo lo que se quiere separar.
LILA = "E6E0F0"
COLOR_CAUSA = {"RC00": LILA, "RC01": AZUL, "RC02": AMBAR, "RC03": ROJO,
               "RC04": AMBAR, "RC05": ROJO, "RC06": ROJO, "RC99": GRIS}

NEGRITA = Font(bold=True)
TITULO = Font(bold=True, size=14)
CHICA = Font(size=9, italic=True, color="595959")
BORDE = Border(*(Side(style="thin", color="BFBFBF"),) * 4)
WRAP = Alignment(wrap_text=True, vertical="top")


# ===========================================================================
# 1. LECTURA
# ===========================================================================

# El reporte de CEDIS lista los SKU con existencia y omite los que están en
# cero, así que en un día que sí se extrajo, la falta de fila ES un cero
# confirmado y no un dato ausente. Confirmado con La Comer el 2026-08-06.
#
# Es la única excepción a "vacío no es cero" en todo el modelo, y por eso está
# aquí arriba y no enterrada en la derivación: sostiene las prioridades 5 a 8
# (separar "CEDIS no tenía" de "CEDIS tenía y no lo mandó"). Sin ella, el
# layout V5 clasifica el 0.3% de los días que llegan a esa rama, porque la
# extracción quitó justo los SKU sin existencia — que son los que explican el
# faltante.
#
# Se aplica SÓLO cuando el día está entre los que la extracción cubre para ese
# CEDIS (ver Fuentes.cedis_cubre). Un día que no se extrajo sigue siendo un
# dato ausente. Poner en False para volver al comportamiento estricto.
CEDIS_AUSENCIA_ES_CERO = True


# ---------------------------------------------------------------------------
# Parche V8 — la existencia de tienda que entrega Tableau es SÓLO la foto de
# CIERRE del día (23:59). No hay mínimo intradía real: la columna "Existencia
# Mínima" llegó duplicando el cierre. Esa foto miente en un día de faltante:
# un producto que llegó tarde, o que es inventario fantasma, aparece con stock
# aunque el anaquel estuvo vacío todo el día. Sin corregirlo, ~80% de los
# faltantes caen en RC01 "Ejecución en tienda" (verificado: 37,752 días, de los
# cuales el 100% tuvo CERO venta ese día).
#
# Criterio: en un día con faltante (OSA<umbral) con existencia de cierre > 0,
# si el SKU NO vendió nada, el cierre no es confiable -> se trata como 0 y el
# día fluye a las ramas de abasto (CEDIS / proveedor) en vez de quedar atrapado
# en RC01. Se PRESERVA RC01 cuando BOPS envió alerta (alerta_enviada = 1): la
# alerta es evidencia de que el producto estaba en tienda y se avisó surtirlo,
# que es ejecución real.
#
# Es un criterio determinista y auditable, como CEDIS_AUSENCIA_ES_CERO. Cuando
# Tableau entregue un mínimo intradía REAL, poner en False: el pipeline ya
# prefiere existencia_minima_dia sobre la foto de cierre y este parche sobra.
INVENTARIO_CIERRE_NO_CONFIABLE = True


# Errores de Excel que llegan como texto en la celda. Aparecen cuando el layout
# trae columnas calculadas con fórmula en vez de dato: si la fórmula no resuelve
# —XLOOKUP guardado como `_xlfn.XLOOKUP` en un Excel que no la reconoce, un
# folio que no existe en la hoja destino— lo que se guarda en caché es el error.
# Se leen como dato ausente, no como valor: un `#N/A` significa que no se supo,
# igual que una celda vacía. Se cuentan al leer para que salga en advertencias.
ERRORES_EXCEL = frozenset({
    "#N/A", "#VALUE!", "#REF!", "#DIV/0!", "#NUM!",
    "#NAME?", "#NULL!", "#SPILL!", "#CALC!", "#GETTING_DATA",
})


def _es_vacio(v) -> bool:
    """Nada, cadena vacía o un error de Excel: los tres son dato ausente."""
    if v is None or v == "":
        return True
    return isinstance(v, str) and v.strip().upper() in ERRORES_EXCEL


def _fecha(v) -> Optional[date]:
    if _es_vacio(v):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return date.fromisoformat(str(v).strip()[:10])


def _texto(v) -> Optional[str]:
    if _es_vacio(v):
        return None
    s = str(v).strip()
    return s or None


def _entero(v) -> Optional[int]:
    if _es_vacio(v):
        return None
    return int(float(v))


def _decimal(v) -> Optional[float]:
    if _es_vacio(v):
        return None
    return float(v)


def _bool_alerta(v) -> Optional[bool]:
    """Bandera 0/1 de BOPS (alerta_enviada / alerta_ejecutada) a bool.

    Respeta 'vacío no es cero': una celda vacía queda en None y la subcausa
    de RC01 no se refina. Sólo un 0 o un 1 explícitos deciden. Tolera texto
    ("1"/"0"/"Sí"/"No") por si el export cambia de tipo.
    """
    if _es_vacio(v):
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = str(v).strip().lower()
    if s in ("1", "si", "sí", "true", "verdadero", "x"):
        return True
    if s in ("0", "no", "false", "falso"):
        return False
    return None


def _osa_pct(v) -> Optional[float]:
    """BOPS reporta OSA binario por diseño de su sistema: 1 = producto visible
    en anaquel ese día, 0 = no visible. No es un % granular, aunque el layout
    lo declara como Decimal 0-100 (confirmado con el equipo, 2026-08-05). Se
    normaliza aquí a la escala 0-100 para que el umbral y el reporte 'OSA %'
    no necesiten un caso especial."""
    v = _decimal(v)
    if v is None:
        return None
    return v * 100.0


def osa_general(fu: "Fuentes") -> Optional[float]:
    """OSA% real del periodo: sobre TODAS las filas de BOPS_OSA (todo SKU,
    tienda y día leído), no sólo los días con faltante que entran a la
    matriz. Es la foto general antes de meterse a la causa raíz de cada día."""
    valores = [v for v in (_osa_pct(fila.get("osa_pct")) for fila in fu.osa.values())
               if v is not None]
    return round(sum(valores) / len(valores), 1) if valores else None


def osa_alcance(fu: "Fuentes") -> Optional[float]:
    """OSA% del periodo sobre los SKU que sí le tocan al análisis.

    Es el número de portada. `osa_general` promedia TODO lo que entregó BOPS,
    y cuando ese export trae divisiones fuera del alcance —como en el layout
    V5— el promedio se hunde por SKU que el catálogo de la tienda ni conoce:
    74.5% contra 85.9% sobre los mismos datos. Uno mide la disponibilidad, el
    otro mide la extracción.

    Sin catálogo con qué comparar no hay alcance que aislar, y devuelve lo
    mismo que osa_general.
    """
    if fu.vacia("CATALOGO"):
        return osa_general(fu)
    valores = [v for (sku, tienda, _), fila in fu.osa.items()
               if fu.en_alcance(sku, tienda)
               for v in (_osa_pct(fila.get("osa_pct")),) if v is not None]
    return round(sum(valores) / len(valores), 1) if valores else None


def universo_osa(fu: "Fuentes", umbral_osa: float) -> Dict[Tuple[str, str], Tuple[int, int]]:
    """(sku, tienda) -> (días medidos, días visibles) sobre TODO el periodo.

    Es el insumo del OSA por SKU del periodo. Se arma aquí y no en
    orcmm_rca_periodo porque allá sólo llegan los días con faltante: los días
    sanos —que son el numerador— nunca salen de Fuentes.

    "Medidos" es a propósito: BOPS no trae fila de todos los SKU todos los
    días (en Coyoacán son ~31.6 días por SKU en un rango de 43). El
    denominador son los días que alguien midió, no los del calendario;
    afirmar algo de un día sin fila sería inventarlo. Los días sin lectura de
    OSA no cuentan ni arriba ni abajo.
    """
    medidos: Dict[Tuple[str, str], int] = defaultdict(int)
    visibles: Dict[Tuple[str, str], int] = defaultdict(int)
    for (sku, tienda, _), fila in fu.osa.items():
        v = _osa_pct(fila.get("osa_pct"))
        if v is None:
            continue
        medidos[(sku, tienda)] += 1
        if v >= umbral_osa:
            visibles[(sku, tienda)] += 1
    return {k: (n, visibles.get(k, 0)) for k, n in medidos.items()}


def waterfall_osa(fu: "Fuentes", diagnosticos: List[dict]) -> dict:
    """De 100% al OSA real, en puntos porcentuales y por causa.

    El Pareto reparte la VENTA PERDIDA; esto reparte el GAP DE OSA, que es
    otra cosa y responde otra pregunta: no "qué causa costó más dinero" sino
    "cuántos puntos de disponibilidad me quitó cada causa". Un día con
    faltante pesa igual que cualquier otro para el OSA, valga lo que valga.

    El denominador es el universo del alcance —las filas de BOPS_OSA cuyo SKU
    está en el catálogo de esa tienda— porque es el mismo del que sale
    `osa_alcance`. Mezclar universos daría un waterfall que no cierra.
    """
    if fu.vacia("CATALOGO"):
        universo = len(fu.osa)
    else:
        universo = sum(1 for (sku, tienda, _) in fu.osa if fu.en_alcance(sku, tienda))

    osa_real = osa_alcance(fu)
    if not universo or osa_real is None:
        return {"osa_teorico": 100.0, "osa_real": osa_real, "universo_filas": universo,
                "escalones": []}

    dias: Dict[str, int] = defaultdict(int)
    ids: Dict[str, str] = {}
    resp: Dict[str, str] = {}
    for dg in dentro_del_alcance(diagnosticos):
        causa = dg["causa_raiz"]
        dias[causa] += 1
        ids[causa] = dg["root_cause_id"]
        resp[causa] = dg["responsable"]

    escalones = [{
        "root_cause_id": ids[c],
        "causa": c,
        "responsable": resp[c],
        "dias": dias[c],
        "puntos_osa": round(dias[c] / universo * 100, 2),
    } for c in dias]
    escalones.sort(key=lambda e: -e["puntos_osa"])

    return {
        "osa_teorico": 100.0,
        "osa_real": osa_real,
        "universo_filas": universo,
        "escalones": escalones,
    }


def fill_rate_proveedor(fu: "Fuentes") -> dict:
    """El cumplimiento del proveedor en una sola cifra, para la portada.

    Se calcula sobre cajas y no promediando los porcentajes de cada
    proveedor: un proveedor con un pedido de 3 cajas no puede pesar lo mismo
    que uno con 6,865 pedidos. Las tres tasas son las mismas del scorecard,
    agregadas.
    """
    pedidas = confirmadas = entregadas = pedidas_con_cita = 0
    pedidos = citas = sin_cita = 0
    for d in desempeno_proveedores(fu):
        pedidas += d.cajas_pedidas
        confirmadas += d.cajas_confirmadas
        entregadas += d.cajas_entregadas
        pedidas_con_cita += d.cajas_pedidas_con_cita
        pedidos += d.pedidos
        citas += d.citas
        sin_cita += d.pedidos_sin_cita

    def tasa(parte, todo):
        return round(parte / todo * 100, 1) if todo else None

    return {
        "proveedores": len(desempeno_proveedores(fu)),
        "pedidos": pedidos,
        "citas": citas,
        "pedidos_sin_cita": sin_cita,
        "cajas_pedidas": pedidas,
        "cajas_confirmadas": confirmadas,
        "cajas_entregadas": entregadas,
        # Entregado sobre pedido: el fill rate de portada.
        "pct_efectivo": tasa(entregadas, pedidas_con_cita),
        # Entregado sobre lo que el proveedor confirmó al agendar.
        "pct_cumplimiento": tasa(entregadas, confirmadas),
        # Confirmado al agendar sobre lo pedido.
        "pct_confirmado": tasa(confirmadas, pedidas_con_cita),
    }


def nivel_servicio_tienda(fu: "Fuentes") -> dict:
    """Qué tanto de lo que la tienda pidió a CEDIS le llegó, en PIEZAS.

    Confirmado con La Comer: SIMA entrega piezas individuales, no cajas
    (`cantidad_pedida_piezas` / `cantidad_surtida_piezas`). No se convierte a
    cajas: habría que dividir entre el empaque del catálogo y el resultado
    heredaría ese error, cuando el dato de origen ya es exacto.

    Se mide sobre piezas y no promediando pedidos, mismo criterio que
    fill_rate_proveedor: un pedido de 4 piezas no puede pesar igual que uno de
    948. Es la pata que faltaba de la cadena —proveedor→CEDIS ya se medía,
    CEDIS→tienda no—, y con ella se puede decir en qué tramo se cae el surtido.
    """
    pedidas = surtidas = 0
    pedidos = completos = sin_surtir = 0
    for p in fu.pedidos_tienda:
        pide = _entero(p.get("cantidad_pedida_piezas"))
        if pide is None:
            continue          # sin denominador no hay tasa: vacío no es cero
        surte = _entero(p.get("cantidad_surtida_piezas")) or 0
        pedidas += pide
        surtidas += surte
        pedidos += 1
        if surte >= pide:
            completos += 1
        elif surte == 0:
            sin_surtir += 1

    return {
        "pedidos": pedidos,
        "piezas_pedidas": pedidas,
        "piezas_surtidas": surtidas,
        "pedidos_completos": completos,
        "pedidos_sin_surtir": sin_surtir,
        # Piezas surtidas sobre pedidas: el nivel de servicio de portada.
        "nivel_servicio_pct": round(surtidas / pedidas * 100, 1) if pedidas else None,
    }


def leer_hoja(wb, nombre: str, advertencias: List[str]) -> List[dict]:
    """Lee una hoja de captura: encabezados en la fila 3, datos desde la 6."""
    if nombre not in wb.sheetnames:
        advertencias.append(f"La hoja {nombre} no existe en el archivo.")
        return []

    ws = wb[nombre]
    # Se traduce el encabezado al nombre canónico del spec: los equipos a veces
    # entregan la columna con el nombre de negocio ("# de Cita" por folio_cita).
    encabezados = [normalizar_encabezado(nombre, c.value) for c in ws[FILA_ENCABEZADO]]

    presentes = [h for h in encabezados if h]
    faltantes = [c[0] for c in HOJAS[nombre]["campos"]
                 if c[2] and c[0] not in presentes]
    if faltantes:
        advertencias.append(f"{nombre}: faltan columnas obligatorias {', '.join(faltantes)}")

    filas = []
    for fila in ws.iter_rows(min_row=FILA_DATOS, values_only=True):
        if all(v is None or v == "" for v in fila):
            continue
        filas.append({h: v for h, v in zip(encabezados, fila) if h})
    return filas


@dataclass
class Fuentes:
    catalogo: Dict[Tuple[str, str], dict] = field(default_factory=dict)
    inv_tienda: Dict[Tuple[str, str, date], dict] = field(default_factory=dict)
    osa: Dict[Tuple[str, str, date], dict] = field(default_factory=dict)
    ventas: Dict[Tuple[str, str, date], dict] = field(default_factory=dict)
    inv_cedis: Dict[Tuple[str, str, date], dict] = field(default_factory=dict)
    transferencias: List[dict] = field(default_factory=list)
    pedidos_tienda: List[dict] = field(default_factory=list)
    pedidos_prov: List[dict] = field(default_factory=list)
    citas_prov: List[dict] = field(default_factory=list)
    # citas indexadas por (folio del pedido, sku): un pedido puede traer varias
    citas_por_pedido: Dict[Tuple[str, str], List[dict]] = field(default_factory=dict)
    # Días que la extracción de CEDIS_INVENTARIO sí cubre, por CEDIS. Ver
    # CEDIS_AUSENCIA_ES_CERO: sólo dentro de esos días la falta de fila se
    # puede leer como existencia cero.
    dias_cedis: Dict[str, Set[date]] = field(default_factory=dict)

    # Catálogo comercial por (sku, tienda): sección, categoría, subcategoría
    # y marca. NO se usa para clasificar —eso sale de `catalogo`, que es el
    # transaccional— sino para poder filtrar y leer el reporte. Sólo se llena
    # leyendo de Postgres; en el camino por archivo queda vacío y el front
    # simplemente no ofrece esos filtros.
    comercial: Dict[Tuple[str, str], dict] = field(default_factory=dict)

    # Eventos agrupados por la clave con la que se consultan y con las fechas
    # ya convertidas. Ver _indexar_eventos: sin esto cada día con faltante
    # recorre la lista completa de eventos.
    #   transferencias_por[(sku, tienda)]   -> [(generacion, salida, recepcion)]
    #   pedidos_tienda_por[(sku, tienda)]   -> [(pedido, surtido)]
    #   pedidos_prov_por[(sku, cedis)]      -> [(pedido, recibo, compromiso, fila)]
    transferencias_por: Dict[Tuple[str, str], List[Tuple]] = field(default_factory=dict)
    pedidos_tienda_por: Dict[Tuple[str, str], List[Tuple]] = field(default_factory=dict)
    pedidos_prov_por: Dict[Tuple[str, str], List[Tuple]] = field(default_factory=dict)
    # (sku, tienda) del catálogo que SIMA no trae. Vacío cuando la
    # exclusión está apagada o SIMA no llegó. Ver EXCLUIR_SKU_SIN_SIMA.
    sin_dato_sima: Set[Tuple[str, str]] = field(default_factory=set)

    conteo: Dict[str, int] = field(default_factory=dict)
    rango: Dict[str, Tuple[Optional[date], Optional[date]]] = field(default_factory=dict)
    advertencias: List[str] = field(default_factory=list)

    def vacia(self, hoja: str) -> bool:
        return self.conteo.get(hoja, 0) == 0

    def en_alcance(self, sku: str, tienda: str) -> bool:
        """¿Este SKU-tienda le toca al análisis?

        Definición ÚNICA del alcance: la usan el OSA de portada, el waterfall,
        el detalle diario y la cobertura. Antes cada uno preguntaba por su
        cuenta si el SKU estaba en el catálogo, y al agregar un segundo motivo
        de exclusión eso se habría desincronizado — con el Pareto calculado
        sobre un universo y el OSA sobre otro, que es la peor forma de fallar
        aquí porque el waterfall deja de cuadrar sin decirlo.
        """
        if (sku, tienda) not in self.catalogo:
            return False
        return (sku, tienda) not in self.sin_dato_sima

    def cedis_cubre(self, cedis: Optional[str], D: date) -> bool:
        """¿La extracción de CEDIS trae ese día para ese CEDIS?

        Es la condición que hace legítimo interpretar la ausencia como cero:
        si el día no se extrajo, la falta de fila no dice nada. Se compara
        contra los días presentes y no contra el rango, porque una extracción
        con huecos (un domingo que no corrió) haría mentir al rango.
        """
        return bool(cedis) and D in self.dias_cedis.get(cedis, ())


@dataclass
class PaqueteFuentes:
    """Todo lo que hace falta para una corrida: el layout y sus CSV.

    Desde el layout V5, TABLEAU_INV_TIENDA y TABLEAU_VENTAS pasan del millón
    de filas que aguanta una hoja de Excel y se entregan aparte. El paquete
    los junta para que el resto del programa vea una sola fuente.
    """
    xlsx: Path
    csvs: Dict[str, List[Path]] = field(default_factory=dict)
    sueltos: List[Path] = field(default_factory=list)

    @classmethod
    def desde(cls, xlsx: Path, csvs: Optional[Iterable[Path]] = None) -> "PaqueteFuentes":
        por_hoja, sueltos = agrupar_por_hoja(csvs or [])
        return cls(Path(xlsx), por_hoja, sueltos)


def leer_fuentes(paquete, umbral_osa: float = 100.0) -> Fuentes:
    """Lee el layout y sus CSV en una sola estructura.

    Acepta un PaqueteFuentes o, por compatibilidad con las corridas que sólo
    tienen Excel, la ruta del .xlsx a secas.

    `umbral_osa` tiene que llegar hasta aquí —y no sólo a derivar_evidencias
    como antes— porque es lo que define qué llaves vale la pena guardar del
    inventario. Sin ese filtro, el layout V5 no cabe en memoria.
    """
    if not isinstance(paquete, PaqueteFuentes):
        paquete = PaqueteFuentes(Path(paquete))

    # read_only: COMPRAS_PEDIDOS_PROV trae 151 mil filas y CEDIS_INVENTARIO
    # declara un millón; cargarlas en el modo normal de openpyxl multiplica
    # por varias veces la memoria del proceso.
    wb = openpyxl.load_workbook(paquete.xlsx, data_only=True, read_only=True)
    fu = Fuentes()
    adv = fu.advertencias

    for suelto in paquete.sueltos:
        adv.append(f"{suelto.name}: no corresponde a ninguna hoja del layout. "
                   f"No se leyó — revisar el nombre del archivo.")

    llaves = (llaves_con_faltante_ws(wb["BOPS_OSA"], umbral_osa)
              if "BOPS_OSA" in wb.sheetnames else set())

    def registrar(hoja: str, filas: List[dict], campos_fecha: List[str],
                  nota_si_vacia: Optional[str] = None):
        fu.conteo[hoja] = len(filas)
        fechas = [_fecha(f.get(c)) for f in filas for c in campos_fecha]
        fechas = [x for x in fechas if x]
        fu.rango[hoja] = (min(fechas), max(fechas)) if fechas else (None, None)

        # Un error de Excel se lee como dato ausente, pero no es lo mismo que un
        # hueco: significa que la columna viene calculada con fórmula y falló.
        # Se avisa por columna para que se pueda pedir el dato en duro.
        errores = Counter(col for f in filas for col, v in f.items()
                          if isinstance(v, str) and v.strip().upper() in ERRORES_EXCEL)
        for col, n in errores.most_common():
            adv.append(f"{hoja}: la columna '{col}' trae {n:,} celdas con error de "
                       f"Excel (#N/A y similares) — viene calculada con fórmula, "
                       f"no con dato. Se leyeron como vacías.")

        if not filas:
            adv.append(f"{hoja}: hoja VACÍA — " + (
                nota_si_vacia or "todas las reglas que dependen de ella "
                                 "quedarán sin clasificar."))

    filas = leer_hoja(wb, "CATALOGO", adv)
    for f in filas:
        fu.catalogo[(_texto(f["sku"]), _texto(f["tienda"]))] = f
    registrar("CATALOGO", filas, [])

    for hoja, destino, llave3, filtrable in [
        ("TABLEAU_INV_TIENDA", fu.inv_tienda, "tienda", True),
        ("BOPS_OSA", fu.osa, "tienda", False),
        ("TABLEAU_VENTAS", fu.ventas, "tienda", True),
        ("CEDIS_INVENTARIO", fu.inv_cedis, "cedis", False),
    ]:
        rutas = paquete.csvs.get(hoja)
        if rutas and hoja in wb.sheetnames:
            adv.append(f"{hoja}: llegó como CSV y también como hoja del Excel. "
                       f"Se usa la hoja del Excel y se ignora el CSV.")
            rutas = None

        if rutas:
            # Sólo los días que el análisis va a mirar. El resto se descarta
            # sin construirle el dict: son millones de filas.
            rep = ReporteCSV(hoja)
            n = 0
            for f in leer_csv(rutas, hoja, llaves=llaves if filtrable else None, rep=rep):
                destino[(f.get("sku"), f.get(llave3), f.get("fecha"))] = f
                n += 1
            fu.conteo[hoja] = rep.filas_leidas
            fu.rango[hoja] = (rep.desde, rep.hasta)
            for a in rep.advertencias:
                adv.append(a)
            if not rep.filas_leidas:
                adv.append(f"{hoja}: los CSV llegaron VACÍOS.")
            elif not n:
                adv.append(f"{hoja}: se leyeron {rep.filas_leidas:,} filas y ninguna "
                           f"corresponde a un día con faltante. Revisar que sean la "
                           f"misma tienda y el mismo periodo que BOPS_OSA.")
            continue

        filas = leer_hoja(wb, hoja, adv)
        for f in filas:
            d = _fecha(f["fecha"])
            if d:
                destino[(_texto(f["sku"]), _texto(f[llave3]), d)] = f
        registrar(hoja, filas, ["fecha"])

    # Qué días cubre la extracción de CEDIS, por CEDIS. Se arma de lo leído y
    # no del rango declarado: es lo que decide dónde la ausencia de fila vale
    # como cero.
    for _, cedis, d in fu.inv_cedis:
        fu.dias_cedis.setdefault(cedis, set()).add(d)

    fu.transferencias = leer_hoja(wb, "CEDIS_TRANSFERENCIAS", adv)
    registrar("CEDIS_TRANSFERENCIAS", fu.transferencias,
              ["fecha_generacion", "fecha_salida_cedis", "fecha_recepcion_tienda"])

    fu.pedidos_tienda = leer_hoja(wb, "SIMA_PEDIDOS_TIENDA", adv)
    registrar("SIMA_PEDIDOS_TIENDA", fu.pedidos_tienda, ["fecha_pedido", "fecha_surtido"],
              "esperado por ahora: SIMA todavía la está procesando. La prioridad 3 está "
              "apagada a propósito para que el árbol pueda seguir hasta CEDIS y proveedor.")

    fu.pedidos_prov = leer_hoja(wb, "COMPRAS_PEDIDOS_PROV", adv)
    registrar("COMPRAS_PEDIDOS_PROV", fu.pedidos_prov, ["fecha_pedido", "fecha_recibo"])

    fu.citas_prov = leer_hoja(wb, "CITAS_PROV_CEDIS", adv)
    registrar("CITAS_PROV_CEDIS", fu.citas_prov, ["fecha_pedido", "fecha_cita"],
              "la prioridad 8 vuelve a operar sólo con cajas pedidas contra cajas "
              "entregadas. No se pierde cobertura, se pierde el detalle: no se puede "
              "separar 'no agendó cita' de 'no cumplió su cita', y un pedido cuya cita "
              "aún no vence se cuenta como incumplimiento del proveedor.")

    for c in fu.citas_prov:
        fu.citas_por_pedido.setdefault(
            (_texto(c.get("folio")), _texto(c.get("sku"))), []).append(c)

    _indexar_eventos(fu)
    _revisar_cobertura_de_citas(fu)
    _marcar_skus_sin_sima(fu)
    _revisar_cobertura_de_sima(fu)

    aviso = aviso_prioridad_3(fu)
    if aviso:
        fu.advertencias.insert(0, aviso)

    return fu


def _indexar_eventos(fu: Fuentes) -> None:
    """Agrupa los eventos por la clave con la que se consultan, con las fechas
    ya convertidas.

    Las tres derivaciones del día D (tránsito, envío, orden a proveedor)
    recorrían la lista COMPLETA de eventos una vez por cada día con faltante.
    Con el volumen real eso son 30,565 días × (30 mil transferencias × 2 +
    151 mil pedidos) ≈ 6.4 mil millones de vueltas, reparseando además la
    misma fecha millones de veces: horas de corrida. Con volúmenes de
    ejemplo nunca se notó.

    Aquí se paga una sola pasada por evento y cada día toca únicamente los
    suyos. El filtro que aplica cada regla no cambia — sólo deja de recorrer
    lo que de todas formas iba a descartar.

    Las listas originales (fu.transferencias, fu.pedidos_prov, ...) quedan
    intactas: el scorecard de proveedores y los reportes de citas siguen
    leyéndolas tal cual.
    """
    for t in fu.transferencias:
        clave = (_texto(t.get("sku")), _texto(t.get("tienda_destino")))
        fu.transferencias_por.setdefault(clave, []).append((
            _fecha(t.get("fecha_generacion")),
            _fecha(t.get("fecha_salida_cedis")),
            _fecha(t.get("fecha_recepcion_tienda")),
        ))

    for p in fu.pedidos_tienda:
        # `origen` es quién generó el pedido: la tienda, o el proceso central
        # (ORIGEN_CENTRALIZADO). Se indexa tal cual y derivar_pedido_tienda se
        # encarga de mirar las dos llaves.
        clave = (_texto(p.get("sku")), _texto(p.get("origen")))
        fu.pedidos_tienda_por.setdefault(clave, []).append((
            _fecha(p.get("fecha_pedido")),
            _fecha(p.get("fecha_surtido")),
        ))

    for o in fu.pedidos_prov:
        clave = (_texto(o.get("sku")), _texto(o.get("cedis_destino")))
        fu.pedidos_prov_por.setdefault(clave, []).append((
            _fecha(o.get("fecha_pedido")),
            _fecha(o.get("fecha_recibo")),
            _compromiso(o),
            o,
        ))


def aviso_prioridad_3(fu: Fuentes) -> Optional[str]:
    """El texto que tiene que acompañar a cualquier resultado sin prioridad 3.

    Va en la consola y en tres hojas del Excel. Un Pareto en el que RC03 no
    puede aparecer se lee mal si no dice por qué: parece que la tienda nunca
    falla, cuando lo que pasa es que no se le preguntó.
    """
    if EVALUAR_PEDIDO_TIENDA:
        return None

    aviso = ("ANÁLISIS PARCIAL — la prioridad 3 está apagada mientras SIMA entrega los "
             "pedidos de tienda. Ningún día puede salir como RC03 'Pedido de tienda no generado', y "
             "los días que le corresponderían se están repartiendo entre RC04 (CEDIS), "
             "RC05 y RC06 (proveedor). Sirve para leer la rama de abasto; NO para repartir "
             "responsabilidades en firme. Se reactiva con EVALUAR_PEDIDO_TIENDA = True.")

    if not fu.vacia("SIMA_PEDIDOS_TIENDA"):
        n = fu.conteo["SIMA_PEDIDOS_TIENDA"]
        aviso += (f" OJO: SIMA_PEDIDOS_TIENDA ya trae {n} "
                  f"{'registro' if n == 1 else 'registros'} y se están IGNORANDO. "
                  f"Ya se puede prender el interruptor.")
    return aviso


def _marcar_skus_sin_sima(fu: Fuentes) -> None:
    """Llena fu.sin_dato_sima con los SKU del catálogo que SIMA no trae.

    GUARDIA: si SIMA no llegó, no se excluye NADA. Sin este `return` una
    entrega sin la hoja dejaría el catálogo entero fuera del alcance y el
    análisis saldría vacío — el modo de fallo más caro posible, porque no
    truena: simplemente no hay resultados y parece que no hubo faltantes.
    """
    if not EXCLUIR_SKU_SIN_SIMA or fu.vacia("SIMA_PEDIDOS_TIENDA"):
        return
    # Mismo criterio que derivar_pedido_tienda: cuenta el pedido propio de la
    # tienda y también el centralizado, que resurte a todas.
    con_pedido = {s for (s, _) in fu.pedidos_tienda_por}
    fu.sin_dato_sima = {(sku, tienda) for (sku, tienda) in fu.catalogo
                        if sku not in con_pedido}


def resumen_excluidos_sima(fu: Fuentes, umbral_osa: float) -> dict:
    """Cuánto se está dejando fuera por no tener datos de SIMA.

    Es el insumo del letrero. La exclusión sube la cobertura a costa de
    achicar el universo, así que el resultado tiene que poder decir sobre
    cuántos SKU se calculó y cuántos se quedaron fuera. Sin esta cifra a la
    vista, el porcentaje de cobertura no significa nada.
    """
    if not fu.sin_dato_sima:
        return {"skus": 0, "dias_con_faltante": 0, "venta_perdida": 0.0,
                "skus_en_alcance": len(fu.catalogo)}

    dias = 0
    venta = 0.0
    for (sku, tienda, _), fila in fu.osa.items():
        if (sku, tienda) not in fu.sin_dato_sima:
            continue
        osa = _osa_pct(fila.get("osa_pct"))
        if osa is None or osa >= umbral_osa:
            continue
        dias += 1
        venta += _decimal(fila.get("venta_perdida_estimada")) or 0.0

    return {
        "skus": len(fu.sin_dato_sima),
        "dias_con_faltante": dias,
        "venta_perdida": round(venta, 2),
        # El denominador honesto: sobre cuántos SKU del catálogo sí se analizó.
        "skus_en_alcance": len(fu.catalogo) - len(fu.sin_dato_sima),
    }


def _revisar_cobertura_de_sima(fu: Fuentes) -> None:
    """Qué tan completa viene SIMA, medido contra el catálogo de la tienda.

    Es el mismo chequeo que _revisar_cobertura_de_citas hace con las citas, y
    por la misma razón: la ausencia de pedido dictamina (RC03, contra la
    tienda), así que hay que saber ANTES de leer el Pareto si la hoja llegó
    completa. Con la primera entrega real cubría 3.5% del catálogo.
    """
    if fu.vacia("SIMA_PEDIDOS_TIENDA") or fu.vacia("CATALOGO"):
        return

    del_catalogo = set(fu.catalogo)
    if not del_catalogo:
        return

    # Un SKU está cubierto si lo pidió su tienda O si viene por pedido
    # centralizado — las dos formas cuentan, igual que en derivar_pedido_tienda.
    centralizados = {s for (s, o) in fu.pedidos_tienda_por if o == ORIGEN_CENTRALIZADO}
    con_pedido = {(s, t) for (s, t) in fu.pedidos_tienda_por if t != ORIGEN_CENTRALIZADO}
    con_pedido |= {(s, t) for (s, t) in del_catalogo if s in centralizados}
    cubiertos = len(con_pedido & del_catalogo)

    pct = round(cubiertos / len(del_catalogo) * 100, 1)
    if pct >= 50:
        return

    fu.advertencias.append(
        f"SIMA_PEDIDOS_TIENDA: {cubiertos:,} de {len(del_catalogo):,} SKU-tienda del catálogo "
        f"({pct}%) tienen algún pedido en el periodo. Los {len(del_catalogo) - cubiertos:,} "
        f"restantes se dictaminan como RC03 'Pedido de tienda no generado', porque La Comer "
        f"confirmó que la extracción viene completa y por lo tanto no aparecer significa que "
        f"la tienda NO pidió. VERIFICAR que siga siendo cierto en esta entrega: si llegara "
        f"recortada, esos días le cobrarían a la tienda un hueco de extracción y RC03 se "
        f"inflaría sin que nada más lo delate.")


def _revisar_cobertura_de_citas(fu: Fuentes) -> None:
    """La ausencia de cita dictamina contra el proveedor, así que conviene
    saber de antemano qué tan completa viene la hoja.

    Un pedido sin fila en CITAS se lee como 'nunca agendó cita'. Si la hoja
    llegó parcial, eso deja de ser un dictamen y se vuelve un artefacto de la
    extracción: hay que verlo antes de leer el Pareto, no después.
    """
    if fu.vacia("CITAS_PROV_CEDIS") or fu.vacia("COMPRAS_PEDIDOS_PROV"):
        return

    pedidos = {(_texto(o.get("folio")), _texto(o.get("sku"))) for o in fu.pedidos_prov}
    con_cita = set(fu.citas_por_pedido)

    huerfanas = con_cita - pedidos
    if huerfanas:
        fu.advertencias.append(
            f"CITAS_PROV_CEDIS: {len(huerfanas)} citas con folio que no existe en "
            f"COMPRAS_PEDIDOS_PROV (ej: {sorted(f for f, _ in huerfanas)[:3]}). "
            f"No se pueden ligar a un pedido y quedan fuera del análisis.")

    sin_cita = pedidos - con_cita
    if sin_cita:
        pct = round(len(sin_cita) / len(pedidos) * 100, 1)
        fu.advertencias.append(
            f"CITAS_PROV_CEDIS: {len(sin_cita)} de {len(pedidos)} pedidos a proveedor "
            f"({pct}%) no tienen cita. El modelo los dictamina como 'el proveedor nunca "
            f"agendó cita' (RC06). CONFIRMAR con Compras que la hoja trae TODAS las citas "
            f"del periodo antes de usar este Pareto: si la extracción vino parcial, la "
            f"causa es un hueco de captura, no del proveedor.")


# ===========================================================================
# 2. DERIVACIÓN — de eventos con fechas a banderas del día D
# ===========================================================================

def sin_acentos(s: str) -> str:
    """Quita tildes y diacríticos, dejando la letra base (Á→A, ñ→n)."""
    return "".join(c for c in unicodedata.normalize("NFKD", s)
                   if not unicodedata.combining(c))


def clave_catalogo(v) -> str:
    """Llave para emparejar un valor de CATALOGO con el vocabulario del motor.

    Sin acentos y en minúsculas, porque el catálogo real NO los escribe igual
    que el spec: el enum dice "Automático" y el sistema entrega "AUTOMATICO".
    Medido sobre la carga real: 22,604 filas con AUTOMATICO que no empataban
    por la tilde, y como el default manda el día a Tienda/Abasto, el modelo
    le estaba cobrando a la TIENDA lo que le toca a COMPRAS (el resurtido
    automático es de ellos — ver RESPONSABLE_PEDIDO_NO_GENERADO).
    """
    return sin_acentos(str(_texto(v) or "")).strip().lower()


VIAS = {clave_catalogo(v.value): v for v in ViaResurtido}
TIPOS_RESURTIDO = {clave_catalogo(t.value): t for t in TipoResurtido}


def derivar_transito_vigente(fu: Fuentes, sku, tienda, D) -> Optional[bool]:
    """Existe transferencia que ya salió de CEDIS y aún no se recibe en tienda."""
    if fu.vacia("CEDIS_TRANSFERENCIAS"):
        return None
    for _, salida, recepcion in fu.transferencias_por.get((sku, tienda), ()):
        if salida and salida <= D and (recepcion is None or D < recepcion):
            return True
    return False


def derivar_envio_generado(fu: Fuentes, sku, tienda, D) -> Optional[bool]:
    """Existe transferencia generada y todavía no recibida en tienda."""
    if fu.vacia("CEDIS_TRANSFERENCIAS"):
        return None
    for generacion, _, recepcion in fu.transferencias_por.get((sku, tienda), ()):
        if generacion and generacion <= D and (recepcion is None or recepcion > D):
            return True
    return False


def derivar_pedido_tienda(fu: Fuentes, sku, tienda, D) -> Optional[bool]:
    """Existe pedido de tienda abierto al día D.

    La ausencia en SIMA es EVIDENCIA de que no se pidió, no un hueco de datos.
    Confirmado con La Comer (2026-08-20): la extracción viene completa, así que
    un SKU del catálogo que no aparece es un SKU que la tienda no pidió en el
    periodo. Por eso devuelve False y el día cae en RC03.

    Esto estuvo al revés entre el 19 y el 20 de agosto. Con la primera entrega
    de SIMA sólo el 0.7% de los SKU con faltante tenía pedido, y un súper no
    resurte 600 claves al mes: parecía extracción parcial, así que la ausencia
    se leía como "no sé" para no fabricar culpables. La segunda entrega —de
    5,707 filas a 57,024, con los pedidos centralizados incluidos— y la
    confirmación de La Comer cerraron la duda: la hoja sí venía completa; lo
    que faltaba eran los centralizados y el resto del periodo.

    QUÉ IMPLICA, porque no es menor: el modelo ahora AFIRMA "la tienda no
    pidió" para decenas de miles de días. Se sostiene sobre la palabra del
    dueño del dato, no sobre lo que el archivo demuestra por sí solo — si una
    entrega futura llegara recortada, RC03 se inflaría en silencio y le
    cobraría a la tienda un hueco de extracción. _revisar_cobertura_de_sima
    mide justamente eso y avisa.

    None sólo cuando la hoja entera viene vacía: ahí no hay nada que afirmar.
    """
    if fu.vacia("SIMA_PEDIDOS_TIENDA"):
        return None
    # Los que generó la tienda MÁS los centralizados. Un pedido centralizado
    # también resurte a la tienda: la diferencia es quién lo generó, no a quién
    # llega. Si sólo se miraran los de la tienda, un SKU que se pide de forma
    # centralizada saldría siempre como "no pidió" — otra vez culpando a la
    # tienda por algo que no le toca hacer.
    #
    # Un pedido con ORIGEN_CENTRALIZADO cubre a TODAS las tiendas, no a una
    # (confirmado con La Comer, 2026-08-19). Por eso no hace falta guardar a
    # qué sucursal pertenece: la misma fila aplica a todas, y cuando viene
    # repetida en el archivo de dos tiendas el upsert por (folio, sku) la deja
    # igual en vez de duplicarla.
    eventos = list(fu.pedidos_tienda_por.get((sku, tienda), ()))
    eventos += fu.pedidos_tienda_por.get((sku, ORIGEN_CENTRALIZADO), ())
    for pedido, surtido in eventos:
        if pedido and pedido <= D and (surtido is None or surtido > D):
            return True
    return False


@dataclass
class OrdenProveedor:
    """Estado de la orden a proveedor que explica el faltante del día D."""
    existe: Optional[bool] = None
    cajas_pedidas: Optional[int] = None
    cajas_entregadas: Optional[int] = None
    folio: Optional[str] = None
    # Bloque de citas. Todo en None = no hay hoja CITAS y la regla 8 opera
    # como antes, comparando cajas pedidas contra entregadas.
    cita_agendada: Optional[bool] = None
    cita_vencida: Optional[bool] = None
    cajas_confirmadas_cita: Optional[int] = None
    fecha_cita: Optional[date] = None
    folio_cita: Optional[str] = None


def _compromiso(o: dict) -> date:
    """Fecha en la que se esperaba la mercancía. Ordena las órdenes abiertas."""
    return _fecha(o.get("fecha_cita")) or _fecha(o.get("fecha_recibo")) \
        or _fecha(o.get("fecha_pedido")) or date.max


def derivar_orden_proveedor(fu: Fuentes, sku, cedis, D) -> OrdenProveedor:
    """Orden a proveedor vigente al día D, con el estado de su cita.

    Con 33 órdenes traslapadas para un mismo SKU, cualquier día tiene varias
    vigentes. Se elige la de compromiso más próximo: es la entrega que debía
    haber tapado el faltante de hoy, y por lo tanto la que lo explica. El
    criterio es determinista, no depende del orden de las filas del Excel.
    """
    if fu.vacia("COMPRAS_PEDIDOS_PROV"):
        return OrdenProveedor()

    vigentes = [(compromiso, o) for pedido, recibo, compromiso, o
                in fu.pedidos_prov_por.get((sku, cedis), ())
                if pedido and pedido <= D and (recibo is None or recibo > D)]

    if not vigentes:
        return OrdenProveedor(existe=False)

    _, o = min(vigentes, key=lambda x: (x[0], _texto(x[1].get("folio")) or ""))
    folio = _texto(o.get("folio"))
    orden = OrdenProveedor(
        existe=True,
        cajas_pedidas=_entero(o.get("cajas_pedidas")),
        cajas_entregadas=0,   # la orden sigue abierta al día D: aún no hay recibo
        folio=folio,
    )

    if fu.vacia("CITAS_PROV_CEDIS"):
        return orden

    return _aplicar_citas(fu, orden, folio, sku, D)


def _aplicar_citas(fu: Fuentes, orden: OrdenProveedor, folio, sku, D) -> OrdenProveedor:
    """Estado de las citas del pedido al día D.

    Se separan las citas ya vencidas de las que aún no llegan. Sólo las
    vencidas juzgan al proveedor: de ellas salen las cajas que confirmó y las
    que realmente entregó. Una cita futura dice lo contrario, que todavía
    está en tiempo.
    """
    citas = fu.citas_por_pedido.get((folio, sku), [])
    if not citas:
        orden.cita_agendada = False
        return orden

    orden.cita_agendada = True
    vencidas = [c for c in citas if (_fecha(c.get("fecha_cita")) or date.max) <= D]

    if vencidas:
        ultima = max(vencidas, key=lambda c: _fecha(c.get("fecha_cita")) or date.min)
        confirmadas = [_entero(c.get("cajas_confirmadas_cita")) for c in vencidas]
        entregadas = [_entero(c.get("cajas_entregadas")) for c in vencidas]
        orden.cita_vencida = True
        orden.cajas_confirmadas_cita = None if None in confirmadas else sum(confirmadas)
        orden.cajas_entregadas = None if None in entregadas else sum(entregadas)
        orden.fecha_cita = _fecha(ultima.get("fecha_cita"))
        orden.folio_cita = _texto(ultima.get("folio_cita"))
        return orden

    proxima = min(citas, key=lambda c: _fecha(c.get("fecha_cita")) or date.max)
    orden.cita_vencida = False
    orden.fecha_cita = _fecha(proxima.get("fecha_cita"))
    orden.folio_cita = _texto(proxima.get("folio_cita"))
    return orden


def derivar_evidencias(fu: Fuentes, umbral_osa: float) -> List[EvidenciaSKUTienda]:
    """Una evidencia por cada día con faltante (OSA por debajo del umbral)."""
    evidencias = []
    inv_cierre_descartado = 0  # días RC01 evitados: cierre>0 sin venta ni alerta
    sin_catalogo = set()      # (sku, tienda) — se reportan juntos al final
    cedis_derivado = 0        # días con existencia de CEDIS leída como cero
    tipos_desconocidos: Counter = Counter()   # tipo_resurtido fuera del vocabulario

    for (sku, tienda, D), fila_osa in sorted(fu.osa.items(), key=lambda kv: (kv[0][0], kv[0][1], kv[0][2])):
        osa = _osa_pct(fila_osa.get("osa_pct"))
        if osa is None or osa >= umbral_osa:
            continue   # día sin faltante: no entra al análisis

        cat = fu.catalogo.get((sku, tienda))
        if cat is None:
            sin_catalogo.add((sku, tienda))
            via, cedis, tipo_resurtido = None, None, None
        else:
            via = VIAS.get(clave_catalogo(cat.get("via_resurtido")))
            cedis = _texto(cat.get("cedis_surtidor"))
            crudo_tipo = _texto(cat.get("tipo_resurtido"))
            tipo_resurtido = TIPOS_RESURTIDO.get(clave_catalogo(crudo_tipo))
            if crudo_tipo and tipo_resurtido is None:
                # El catálogo trae un valor que el modelo no conoce. NO se
                # adivina a qué responsable equivale: se cuenta y se reporta,
                # para que La Comer diga qué significa cada uno.
                tipos_desconocidos[crudo_tipo] += 1

        inv_t = fu.inv_tienda.get((sku, tienda, D))
        existencia = None
        if inv_t is not None:
            existencia = _entero(inv_t.get("existencia_minima_dia"))
            if existencia is None:
                existencia = _entero(inv_t.get("existencia_piezas"))

        # Parche V8: el cierre miente cuando el SKU no vendió nada ese día.
        # Ver INVENTARIO_CIERRE_NO_CONFIABLE. Se preserva RC01 si BOPS alertó.
        if INVENTARIO_CIERRE_NO_CONFIABLE and existencia is not None and existencia > 0:
            venta_row = fu.ventas.get((sku, tienda, D))
            vendidas = None
            if venta_row is not None:
                vendidas = _decimal(venta_row.get("unidades_vendidas"))
                if vendidas is None:
                    vendidas = _decimal(venta_row.get("importe_venta"))
            alerto = _bool_alerta(fila_osa.get("alerta_enviada"))
            if not vendidas and not alerto:
                existencia = 0
                inv_cierre_descartado += 1

        inv_c = fu.inv_cedis.get((sku, cedis, D)) if cedis else None
        existencia_cedis = None
        if inv_c is not None:
            existencia_cedis = _entero(inv_c.get("existencia_piezas"))
            reservadas = _entero(inv_c.get("piezas_reservadas")) or 0
            existencia_cedis = max(0, existencia_cedis - reservadas)
        elif CEDIS_AUSENCIA_ES_CERO and fu.cedis_cubre(cedis, D):
            # El reporte omite los SKU en cero, así que en un día extraído la
            # ausencia es un cero confirmado. Ver CEDIS_AUSENCIA_ES_CERO.
            existencia_cedis = 0
            cedis_derivado += 1

        # La venta perdida manda desde BOPS_OSA: es quien mide cuánto tiempo
        # estuvo vacío el anaquel, y viene para todos los días con faltante.
        # TABLEAU_VENTAS queda de respaldo para los layouts viejos, donde ese
        # campo todavía vivía ahí.
        venta_perdida = _decimal(fila_osa.get("venta_perdida_estimada"))
        if venta_perdida is None:
            venta = fu.ventas.get((sku, tienda, D))
            venta_perdida = _decimal(venta.get("venta_perdida_estimada")) if venta else None

        orden = derivar_orden_proveedor(fu, sku, cedis, D) if cedis else OrdenProveedor()

        evidencias.append(EvidenciaSKUTienda(
            sku=sku, tienda=tienda, fecha=D,
            osa=osa,
            venta_perdida=venta_perdida,
            # None si no hay catálogo con qué comprobarlo: sin la hoja no se
            # puede afirmar que un SKU esté fuera del alcance.
            en_catalogo=None if fu.vacia("CATALOGO") else cat is not None,
            # None cuando no se evalúa (SIMA vacía o exclusión apagada):
            # así la regla de alcance no dispara y nada se excluye.
            sku_en_sima=(None if not fu.sin_dato_sima
                         else (sku, tienda) not in fu.sin_dato_sima),
            inventario_tienda=existencia,
            # Banderas de alerta de BOPS (V8). Refinan la subcausa de RC01;
            # None = sin dato (vacío no es cero).
            alerta_enviada=_bool_alerta(fila_osa.get("alerta_enviada")),
            alerta_ejecutada=_bool_alerta(fila_osa.get("alerta_ejecutada")),
            transito_vigente=derivar_transito_vigente(fu, sku, tienda, D),
            pedido_tienda_generado=derivar_pedido_tienda(fu, sku, tienda, D),
            tipo_resurtido=tipo_resurtido,
            via_resurtido=via,
            inventario_cedis=existencia_cedis,
            envio_cedis_generado=derivar_envio_generado(fu, sku, tienda, D),
            pedido_proveedor_generado=orden.existe,
            proveedor_cajas_pedidas=orden.cajas_pedidas,
            proveedor_cajas_entregadas=orden.cajas_entregadas,
            proveedor_cita_agendada=orden.cita_agendada,
            proveedor_cita_vencida=orden.cita_vencida,
            proveedor_cajas_confirmadas_cita=orden.cajas_confirmadas_cita,
            proveedor_fecha_cita=orden.fecha_cita,
            proveedor_folio_pedido=orden.folio,
            proveedor_folio_cita=orden.folio_cita,
        ))

    # Una advertencia por motivo, no por día: con un catálogo parcial esto
    # serían decenas de miles de líneas repetidas en el Excel y en el JSON.
    if sin_catalogo:
        skus = sorted({s for s, _ in sin_catalogo})
        n = len(sin_catalogo)
        fu.advertencias.append(
            f"BOPS_OSA trae {n} {'combinación' if n == 1 else 'combinaciones'} "
            f"SKU-tienda con faltante que no existen en CATALOGO ({len(skus)} "
            f"{'SKU distinto' if len(skus) == 1 else 'SKU distintos'}, ej: "
            f"{', '.join(skus[:5])}). Se cuentan como FUERA DEL ALCANCE, no como "
            f"dato faltante: el catálogo define qué SKU le tocan a este análisis. "
            f"Si la extracción de OSA debería venir filtrada a la misma división "
            f"que el catálogo, hay que pedirla así.")

    if cedis_derivado:
        fu.advertencias.append(
            f"CEDIS_INVENTARIO: {cedis_derivado} días se resolvieron con existencia "
            f"CERO derivada de la ausencia de fila, porque el reporte omite los SKU "
            f"sin existencia (confirmado con La Comer). Es el único lugar del modelo "
            f"donde vacío se lee como cero; se apaga con CEDIS_AUSENCIA_ES_CERO.")

    if tipos_desconocidos:
        detalle = ", ".join(f"{v} ({n:,} días)" for v, n in tipos_desconocidos.most_common(6))
        fu.advertencias.append(
            f"CATALOGO: {len(tipos_desconocidos)} valores de tipo_resurtido que el modelo "
            f"no conoce — {detalle}. El vocabulario de la matriz es sólo Manual y "
            f"Automático, así que estos días NO afinan el responsable de RC03 y caen al "
            f"default 'Tienda / Abasto'. CONFIRMAR con La Comer a qué responsable "
            f"equivale cada uno: si alguno es resurtido automático, hoy se le está "
            f"cobrando a la tienda algo que le toca a Compras.")

    if inv_cierre_descartado:
        fu.advertencias.append(
            f"TABLEAU_INV_TIENDA: {inv_cierre_descartado} días con faltante traían "
            f"existencia de CIERRE > 0 pero CERO venta y sin alerta de BOPS, así que "
            f"el inventario de cierre se descartó (tratado como 0) para que el día "
            f"fluya a las ramas de abasto en vez de caer en RC01 'Ejecución en "
            f"tienda'. Tableau sólo entrega la foto de cierre (23:59), no el mínimo "
            f"intradía; se apaga con INVENTARIO_CIERRE_NO_CONFIABLE.")

    return evidencias


# ===========================================================================
# 3. SALIDA
# ===========================================================================

def si_no(v: Optional[bool]) -> str:
    return "" if v is None else ("Sí" if v else "No")


def _celda_cita(ev: EvidenciaSKUTienda) -> str:
    """Resume en una celda si el pedido tenía cita y si ya había vencido."""
    if ev.proveedor_cita_agendada is None:
        return ""
    if not ev.proveedor_cita_agendada:
        return "Sin cita"
    fecha = ev.proveedor_fecha_cita.isoformat() if ev.proveedor_fecha_cita else "s/f"
    return f"{fecha} {'vencida' if ev.proveedor_cita_vencida else 'por vencer'}"


def _banner(ws, fila: int, col_ini: int, col_fin: int, texto: str, alto: int = 46):
    """Aviso a todo lo ancho, para que no se pueda leer la tabla sin verlo."""
    c = ws.cell(row=fila, column=col_ini, value=texto)
    c.font = Font(bold=True, size=9, color="9C0006")
    c.fill = PatternFill("solid", fgColor=ROJO)
    c.alignment = WRAP
    c.border = BORDE
    ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
    ws.row_dimensions[fila].height = alto


def _encabezado(ws, fila, valores):
    for j, v in enumerate(valores, 1):
        c = ws.cell(row=fila, column=j, value=v)
        c.font = NEGRITA
        c.fill = PatternFill("solid", fgColor=AMARILLO)
        c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        c.border = BORDE


@dataclass
class DesempenoProveedor:
    """Cumplimiento de un proveedor contra CEDIS en el periodo.

    Las tres tasas de la cita se miden sobre el mismo universo, los pedidos que
    SÍ tienen cita, para que se puedan leer en cadena:
      confirmado   = de lo que se le pidió, ¿a cuánto se comprometió al agendar?
      cumplimiento = de lo que confirmó, ¿cuánto entregó?
      efectivo     = de lo que se le pidió, ¿cuánto llegó?  (el que duele)

    Los pedidos sin cita se reportan aparte y no se meten al denominador: hoy
    no se sabe si el proveedor nunca agendó o si la extracción vino corta, y
    mezclarlos hundiría la tasa por una razón que todavía no está confirmada.
    El surtido del pedido (columna aparte) sí cubre los 33 pedidos, tomado de
    COMPRAS_PEDIDOS_PROV.
    """
    proveedor_id: str
    nombre: str
    # Los IDs que se consolidaron en este renglón. Casi siempre es uno; cuando
    # son varios, el mismo proveedor viene dado de alta más de una vez en
    # COMPRAS (ver _clave_proveedor). Se conservan para poder rastrearlo.
    ids: List[str] = field(default_factory=list)
    # Disponibilidad de los SKU de este proveedor en el periodo. None cuando
    # ninguno de sus SKU está en el catálogo de la tienda analizada.
    osa_periodo: Optional[float] = None
    dias_evaluados: Optional[int] = None
    pedidos: int = 0
    cajas_pedidas: int = 0
    cajas_surtidas_pedido: int = 0     # cierre de la orden, según COMPRAS
    pedidos_con_cita: int = 0
    cajas_pedidas_con_cita: int = 0
    citas: int = 0
    citas_incumplidas: int = 0
    cajas_confirmadas: int = 0
    cajas_entregadas: int = 0

    @property
    def pedidos_sin_cita(self) -> int:
        return self.pedidos - self.pedidos_con_cita

    @staticmethod
    def _tasa(parte: int, todo: int) -> Optional[float]:
        return round(parte / todo, 4) if todo else None

    @property
    def pct_surtido_pedido(self) -> Optional[float]:
        return self._tasa(self.cajas_surtidas_pedido, self.cajas_pedidas)

    @property
    def pct_confirmado(self) -> Optional[float]:
        return self._tasa(self.cajas_confirmadas, self.cajas_pedidas_con_cita)

    @property
    def pct_cumplimiento(self) -> Optional[float]:
        return self._tasa(self.cajas_entregadas, self.cajas_confirmadas)

    @property
    def pct_efectivo(self) -> Optional[float]:
        return self._tasa(self.cajas_entregadas, self.cajas_pedidas_con_cita)


def universo_osa_por_proveedor(fu: "Fuentes", umbral_osa: float
                                ) -> Dict[str, Tuple[int, int]]:
    """proveedor_id -> (días medidos, días visibles) de TODOS sus SKU.

    Insumo del "OSA del proveedor". Se indexa por ID y no por nombre a
    propósito: el catálogo escribe "MARCAS NESTLE" y compras "MARCAS NESTLE,
    S.A. DE C.V.", así que emparejar por nombre fallaría justo en los
    proveedores grandes. El ID sí calza exacto entre las dos fuentes.

    OJO con lo que mide: es la disponibilidad de los productos de ese
    proveedor, NO su culpa. Un día que el SKU faltó por ejecución en tienda
    (RC01) también baja este número. Es descriptivo, no atributivo.
    """
    medidos: Dict[str, int] = defaultdict(int)
    visibles: Dict[str, int] = defaultdict(int)
    for (sku, tienda), (m, v) in universo_osa(fu, umbral_osa).items():
        cat = fu.catalogo.get((sku, tienda))
        if not cat:
            continue
        pid = _texto(cat.get("proveedor_id"))
        if not pid:
            continue
        medidos[pid] += m
        visibles[pid] += v
    return {p: (n, visibles.get(p, 0)) for p, n in medidos.items()}


def _clave_proveedor(nombre: Optional[str], pid: Optional[str]) -> str:
    """Clave para consolidar un proveedor que viene dado de alta varias veces.

    Verificado contra los datos reales de COMPRAS: el mismo proveedor aparece
    con dos IDs distintos —Nestlé como 16661 "MARCAS NESTLE, S.A. DE C.V." y
    16653 "MARCAS NESTLE SA DE CV", PepsiCo como 154833 y 929482— así que
    agrupar por ID lo parte en dos renglones que compiten entre sí en el
    ranking. Y al revés: 8 IDs traen el mismo nombre escrito de dos formas que
    sólo difieren en espacios dobles ("SA DE  CV").

    Por eso la clave es el NOMBRE normalizado y el ID sólo se usa cuando no hay
    nombre. Los IDs originales no se pierden: van en `ids`.

    La normalización se queda SÓLO con letras y números: quitar la puntuación
    no basta, porque "S.A. DE C.V." se vuelve "S A DE C V" y no empata con
    "SA DE CV" —que es justo el caso de Nestlé—. Sin los espacios, las dos
    grafías caen en la misma clave.
    """
    n = _texto(nombre)
    if not n:
        return f"#{_texto(pid) or '(sin id)'}"
    return "".join(c for c in n.upper() if c.isalnum())


def desempeno_proveedores(fu: Fuentes, umbral_osa: float = 100.0
                           ) -> List[DesempenoProveedor]:
    """Scorecard por proveedor, construido sobre las hojas de origen.

    No depende de la clasificación: mide al proveedor sobre TODOS sus pedidos
    del periodo, no sólo sobre los días en que hubo faltante en tienda. Un
    proveedor puede incumplir sin que se note en anaquel, y eso también hay
    que verlo antes de que sí se note.
    """
    prov: Dict[str, DesempenoProveedor] = {}

    for o in fu.pedidos_prov:
        pid = _texto(o.get("proveedor_id")) or "(sin id)"
        nombre = _texto(o.get("proveedor_nombre")) or ""
        clave = _clave_proveedor(nombre, pid)
        d = prov.setdefault(clave, DesempenoProveedor(pid, nombre))
        if pid not in d.ids:
            d.ids.append(pid)
        # Se queda la grafía más larga: entre "MARCAS NESTLE SA DE CV" y
        # "MARCAS NESTLE, S.A. DE C.V." la segunda es la que trae la razón
        # social completa, y es la que una persona reconoce.
        if len(nombre) > len(d.nombre):
            d.nombre = nombre
        pedidas = _entero(o.get("cajas_pedidas")) or 0
        d.pedidos += 1
        d.cajas_pedidas += pedidas
        d.cajas_surtidas_pedido += _entero(o.get("cajas_entregadas")) or 0

        citas = fu.citas_por_pedido.get((_texto(o.get("folio")), _texto(o.get("sku"))), [])
        if not citas:
            continue
        d.pedidos_con_cita += 1
        d.cajas_pedidas_con_cita += pedidas
        for c in citas:
            confirmadas = _entero(c.get("cajas_confirmadas_cita")) or 0
            entregadas = _entero(c.get("cajas_entregadas")) or 0
            d.citas += 1
            d.cajas_confirmadas += confirmadas
            d.cajas_entregadas += entregadas
            if entregadas < confirmadas:
                d.citas_incumplidas += 1

    # OSA de los SKU de cada proveedor. Se suma sobre TODOS los IDs que se
    # consolidaron en el renglón: Nestlé viene con dos altas distintas y su
    # universo de SKU está repartido entre las dos.
    universo = universo_osa_por_proveedor(fu, umbral_osa)
    for d in prov.values():
        medidos = sum(universo.get(p, (0, 0))[0] for p in d.ids)
        visibles = sum(universo.get(p, (0, 0))[1] for p in d.ids)
        if medidos:
            d.dias_evaluados = medidos
            d.osa_periodo = round(visibles / medidos * 100, 1)

    return sorted(prov.values(), key=lambda d: d.cajas_pedidas, reverse=True)


def citas_incumplidas(fu: Fuentes) -> List[dict]:
    """Citas vencidas en las que el proveedor entregó menos de lo confirmado."""
    fallas = []
    for c in fu.citas_prov:
        confirmadas = _entero(c.get("cajas_confirmadas_cita"))
        entregadas = _entero(c.get("cajas_entregadas"))
        if confirmadas is None or entregadas is None or entregadas >= confirmadas:
            continue
        fallas.append({
            "folio": _texto(c.get("folio")),
            "folio_cita": _texto(c.get("folio_cita")),
            "sku": _texto(c.get("sku")),
            "proveedor": _texto(c.get("proveedor_nombre")) or _texto(c.get("proveedor_id")),
            "fecha_cita": _fecha(c.get("fecha_cita")),
            "confirmadas": confirmadas,
            "entregadas": entregadas,
            "faltantes": confirmadas - entregadas,
            "estatus": _texto(c.get("estatus_cita")) or "",
        })
    fallas.sort(key=lambda f: f["faltantes"], reverse=True)
    return fallas


def discrepancias_pedido_cita(fu: Fuentes) -> List[dict]:
    """Dónde el pedido y la cita cuentan historias distintas del mismo folio.

    Las dos hojas traen fecha_cita y cajas entregadas. Cuando no coinciden hay
    que decidir cuál manda antes de presentar cualquier número al proveedor:
    aquí manda la cita para lo que pasó en la cita, y el pedido para el cierre
    de la orden. Esta lista es para que Compras concilie el origen del dato.
    """
    salida = []
    for o in fu.pedidos_prov:
        folio, sku = _texto(o.get("folio")), _texto(o.get("sku"))
        citas = fu.citas_por_pedido.get((folio, sku), [])
        if not citas:
            continue

        ent_pedido = _entero(o.get("cajas_entregadas"))
        ent_cita = sum(_entero(c.get("cajas_entregadas")) or 0 for c in citas)
        cita_pedido = _fecha(o.get("fecha_cita"))
        fechas_cita = sorted(f for f in (_fecha(c.get("fecha_cita")) for c in citas) if f)

        motivos = []
        if ent_pedido is not None and ent_pedido != ent_cita:
            motivos.append(f"cajas entregadas: pedido dice {ent_pedido}, citas suman {ent_cita}")
        if cita_pedido and fechas_cita and cita_pedido not in fechas_cita:
            motivos.append(f"fecha de cita: pedido dice {cita_pedido.isoformat()}, "
                           f"cita dice {fechas_cita[0].isoformat()}")
        if motivos:
            salida.append({"folio": folio, "sku": sku, "motivos": " · ".join(motivos)})
    return salida


def escribir_hoja_proveedor(wb, fu: Fuentes, diagnosticos: List[dict]) -> None:
    ws = wb.create_sheet("Proveedor y citas")
    ws.sheet_view.showGridLines = False
    for j, w in enumerate([3, 30, 9, 11, 13, 8, 10, 13, 11, 12, 13, 13, 12, 13], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws["B2"] = "Cumplimiento del proveedor en CEDIS"
    ws["B2"].font = TITULO
    ws["B3"] = ("El pedido dice qué se le pidió; la cita dice a qué se comprometió y qué entregó. "
                "Las tres tasas de cita se miden sobre los pedidos que tienen cita: confirmado = "
                "cajas de cita sobre cajas pedidas, cumplimiento = entregado sobre confirmado, "
                "efectivo = entregado sobre pedido. Los pedidos sin cita van aparte.")
    ws["B3"].font = CHICA

    if fu.vacia("CITAS_PROV_CEDIS"):
        ws["B5"] = ("CITAS_PROV_CEDIS viene vacía. Sin ella no se puede separar 'no agendó cita' "
                    "de 'no cumplió su cita', y la prioridad 8 vuelve a operar sólo con "
                    "cajas pedidas contra cajas entregadas.")
        ws["B5"].fill = PatternFill("solid", fgColor=AMBAR)
        ws["B5"].alignment = WRAP
        ws.merge_cells(start_row=5, start_column=2, end_row=5, end_column=14)
        return

    f = 5
    _encabezado(ws, f, ["", "Proveedor", "Pedidos", "Cajas ped.", "% surtido pedido",
                        "Citas", "Sin cita", "Cajas ped. c/cita", "Cajas conf.",
                        "Cajas entreg.", "% confirmado", "% cumplim.", "% efectivo",
                        "Citas falladas"])
    ws.row_dimensions[f].height = 44
    f += 1

    for d in desempeno_proveedores(fu):
        etiqueta = f"{d.nombre} ({d.proveedor_id})" if d.nombre else d.proveedor_id
        valores = [etiqueta, d.pedidos, d.cajas_pedidas, d.pct_surtido_pedido,
                   d.citas, d.pedidos_sin_cita, d.cajas_pedidas_con_cita,
                   d.cajas_confirmadas, d.cajas_entregadas,
                   d.pct_confirmado, d.pct_cumplimiento, d.pct_efectivo,
                   d.citas_incumplidas]
        for j, v in enumerate(valores, 2):
            c = ws.cell(row=f, column=j, value=v)
            c.border = BORDE
        for col in (5, 11, 12, 13):
            ws.cell(row=f, column=col).number_format = "0.0%"
        for col, tasa in ((5, d.pct_surtido_pedido), (13, d.pct_efectivo)):
            if tasa is not None:
                ws.cell(row=f, column=col).fill = PatternFill(
                    "solid", fgColor=VERDE if tasa >= 0.95 else
                    (AMBAR if tasa >= 0.8 else ROJO))
        if d.pedidos_sin_cita:
            ws.cell(row=f, column=7).fill = PatternFill("solid", fgColor=AMBAR)
        f += 1

    # --- Impacto de cada subcausa (pedido de tienda y proveedor) -----------
    subcausas = resumen_por_subcausa(diagnosticos)

    f += 2
    ws.cell(row=f, column=2, value="Qué detalle de la causa costó más").font = TITULO
    f += 1
    ws.cell(row=f, column=2, value=(
        "Sólo los días con faltante que la matriz alcanzó a clasificar en la prioridad 3 "
        "(pedido de tienda, según CATALOGO.tipo_resurtido) u 8 (proveedor). Si está vacío, "
        "el árbol se detuvo antes de llegar a alguna de las dos.")).font = CHICA
    f += 2

    _encabezado(ws, f, ["", "Detalle / subcausa", "Días", "Venta perdida"] + [""] * 10)
    f += 1
    if subcausas:
        for fila in subcausas:
            ws.cell(row=f, column=2, value=fila["subcausa"]).alignment = WRAP
            ws.cell(row=f, column=3, value=fila["dias"])
            ws.cell(row=f, column=4, value=fila["venta_perdida"]).number_format = '$#,##0.00'
            for j in range(2, 5):
                ws.cell(row=f, column=j).border = BORDE
            f += 1
    else:
        c = ws.cell(row=f, column=2, value="Ningún día llegó a la prioridad 3 u 8 con dictamen de subcausa.")
        c.fill = PatternFill("solid", fgColor=GRIS)
        f += 1

    # --- Citas falladas, una por una --------------------------------------
    fallas = citas_incumplidas(fu)
    f += 2
    ws.cell(row=f, column=2, value="Citas en las que el proveedor entregó de menos").font = TITULO
    f += 2
    _encabezado(ws, f, ["", "Proveedor", "Pedido", "Cita", "Fecha cita", "SKU",
                        "Confirmó", "Entregó", "Faltaron", "Estatus del sistema"] + [""] * 4)
    f += 1
    for x in fallas:
        valores = [x["proveedor"], x["folio"], x["folio_cita"],
                   x["fecha_cita"].isoformat() if x["fecha_cita"] else "",
                   x["sku"], x["confirmadas"], x["entregadas"], x["faltantes"], x["estatus"]]
        for j, v in enumerate(valores, 2):
            c = ws.cell(row=f, column=j, value=v)
            c.border = BORDE
        ws.cell(row=f, column=10).fill = PatternFill("solid", fgColor=ROJO)
        f += 1
    if not fallas:
        ws.cell(row=f, column=2, value="Ninguna: el proveedor entregó todo lo que confirmó.")\
            .fill = PatternFill("solid", fgColor=VERDE)
        f += 1

    # --- Conciliación entre las dos hojas ---------------------------------
    disc = discrepancias_pedido_cita(fu)
    if disc:
        f += 2
        ws.cell(row=f, column=2, value="Pedido y cita no cuadran").font = TITULO
        f += 1
        ws.cell(row=f, column=2, value=(
            "El mismo folio trae dos versiones del mismo dato. Conciliar con Compras antes "
            "de presentarle números al proveedor.")).font = CHICA
        f += 2
        _encabezado(ws, f, ["", "Pedido", "SKU", "Qué no cuadra"] + [""] * 10)
        f += 1
        # Sin merge_cells por fila: openpyxl compara cada rango fusionado
        # contra todos los anteriores, y con decenas de miles de filas eso
        # deja de terminar. El texto se desborda sobre las columnas vacías de
        # la derecha, que se lee igual y escribe en un instante.
        ambar = PatternFill("solid", fgColor=AMBAR)
        for d in disc:
            ws.cell(row=f, column=2, value=d["folio"]).border = BORDE
            ws.cell(row=f, column=3, value=d["sku"]).border = BORDE
            ws.cell(row=f, column=4, value=d["motivos"]).fill = ambar
            f += 1


def escribir_resultado(ruta: Path, fu: Fuentes, evidencias: List[EvidenciaSKUTienda],
                       diagnosticos: List[dict], umbral_osa: float = 100.0):
    wb = openpyxl.Workbook()
    aviso = aviso_prioridad_3(fu)

    # El Pareto y el detalle por SKU se calculan sobre el alcance; la
    # clasificación diaria y la cobertura, sobre todo lo que entregó BOPS.
    # Ver dentro_del_alcance().
    en_alcance = dentro_del_alcance(diagnosticos)

    # --- Clasificación diaria --------------------------------------------
    ws = wb.active
    ws.title = "Clasificacion diaria"
    ws.sheet_view.showGridLines = False
    for j, w in enumerate([16, 9, 12, 8, 13, 11, 10, 11, 12, 8, 11, 11, 10, 10, 12, 11, 10,
                           8, 26, 20, 30, 10, 26, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws["A1"] = "Clasificación de desabasto por causa raíz — resultado por día"
    ws["A1"].font = TITULO
    ws["A2"] = ("Las columnas de evidencia se derivaron de las hojas de eventos; la causa la dictaminó "
                "la matriz RC01-RC06. Un día Sin clasificar nombra el dato que faltó.")
    ws["A2"].font = CHICA
    if aviso:
        _banner(ws, 3, 1, 24, aviso, alto=32)

    _encabezado(ws, 4, [
        "sku", "tienda", "fecha", "OSA %", "venta perdida",
        "inv tienda", "tránsito", "pedido tda", "tipo resurtido", "vía", "inv CEDIS",
        "envío CEDIS", "ped prov", "cajas ped", "cita prov", "cajas conf", "cajas ent",
        "RC", "causa raíz", "responsable", "detalle / subcausa",
        "prioridad", "fuente", "evidencia / dato faltante",
    ])
    ws.row_dimensions[4].height = 32

    for i, (ev, dg) in enumerate(zip(evidencias, diagnosticos)):
        r = 5 + i
        detalle = dg["evidencia"][-1] if dg["evidencia"] else ""
        if dg["datos_faltantes"]:
            detalle = "FALTA EL DATO: " + ", ".join(dg["datos_faltantes"])

        valores = [ev.sku, ev.tienda, ev.fecha.isoformat(), ev.osa, ev.venta_perdida,
                   ev.inventario_tienda, si_no(ev.transito_vigente),
                   si_no(ev.pedido_tienda_generado),
                   ev.tipo_resurtido.value if ev.tipo_resurtido else "",
                   ev.via_resurtido.value if ev.via_resurtido else "",
                   ev.inventario_cedis, si_no(ev.envio_cedis_generado),
                   si_no(ev.pedido_proveedor_generado),
                   ev.proveedor_cajas_pedidas, _celda_cita(ev),
                   ev.proveedor_cajas_confirmadas_cita, ev.proveedor_cajas_entregadas,
                   dg["root_cause_id"], dg["causa_raiz"], dg["responsable"],
                   dg.get("subcausa") or "",
                   dg["prioridad_regla"], dg["fuente"], detalle]

        for j, v in enumerate(valores, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDE
        rc = ws.cell(row=r, column=18)
        rc.font = NEGRITA
        rc.fill = PatternFill("solid", fgColor=COLOR_CAUSA.get(dg["root_cause_id"], GRIS))
        rc.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "D5"
    ws.auto_filter.ref = f"A4:X{4 + len(evidencias)}"

    # --- Pareto -----------------------------------------------------------
    ws = wb.create_sheet("Pareto")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [4, 40, 12, 18, 14, 26]):
        ws.column_dimensions[col].width = w

    diags = diagnosticar_periodo(en_alcance, universo_osa(fu, umbral_osa))
    par = pareto_periodo(diags)

    ws["B2"] = "Pareto del periodo"
    ws["B2"].font = TITULO
    ws["B3"] = (f"{par['sku_tienda_analizados']} combinaciones SKU-tienda · "
                f"{par['dias_con_faltante']} días con faltante · "
                f"venta perdida ${par['venta_perdida_total']:,.2f}")
    ws["B3"].font = CHICA

    f = 5
    if aviso:
        _banner(ws, f, 2, 6, aviso, alto=60)
        f += 2

    # De 100% al OSA real, en puntos. Responde otra pregunta que el Pareto de
    # pesos: no qué causa costó más dinero, sino cuántos puntos de
    # disponibilidad quitó cada una. Un día con faltante pesa igual que
    # cualquier otro para el OSA, valga lo que valga.
    agua = waterfall_osa(fu, diagnosticos)
    if agua["escalones"]:
        ws.cell(row=f, column=2, value="Dónde se pierde la disponibilidad").font = TITULO
        f += 1
        ws.cell(row=f, column=2, value=(
            f"Sobre las {agua['universo_filas']:,} filas de BOPS_OSA cuyo SKU está en "
            f"el catálogo de la tienda. Los puntos suman el gap entre el 100% teórico "
            f"y el OSA real.")).font = CHICA
        f += 2
        _encabezado(ws, f, ["", "Escalón", "Días", "Puntos de OSA", "", "Responsable"])
        f += 1
        ws.cell(row=f, column=2, value="OSA teórico").font = NEGRITA
        ws.cell(row=f, column=4, value=agua["osa_teorico"] / 100).number_format = "0.0%"
        for j in range(2, 7):
            ws.cell(row=f, column=j).border = BORDE
        f += 1
        for e in agua["escalones"]:
            ws.cell(row=f, column=2, value=f"{e['root_cause_id']} · {e['causa']}").fill = \
                PatternFill("solid", fgColor=COLOR_CAUSA.get(e["root_cause_id"], GRIS))
            ws.cell(row=f, column=3, value=e["dias"])
            ws.cell(row=f, column=4, value=-e["puntos_osa"] / 100).number_format = "0.00%"
            ws.cell(row=f, column=6, value=e["responsable"])
            for j in range(2, 7):
                ws.cell(row=f, column=j).border = BORDE
            f += 1
        ws.cell(row=f, column=2, value="OSA real del periodo").font = NEGRITA
        c = ws.cell(row=f, column=4, value=(agua["osa_real"] or 0) / 100)
        c.number_format = "0.0%"
        c.font = NEGRITA
        for j in range(2, 7):
            ws.cell(row=f, column=j).border = BORDE
            ws.cell(row=f, column=j).fill = PatternFill("solid", fgColor=AMARILLO)
        f += 3

    _encabezado(ws, f, ["", "Causa raíz", "Días", "Venta perdida", "% del total", "Responsable"])
    f += 1
    for fila in resumen_por_causa(en_alcance):
        ws.cell(row=f, column=2, value=f"{fila['root_cause_id']} · {fila['causa']}")
        ws.cell(row=f, column=2).fill = PatternFill(
            "solid", fgColor=COLOR_CAUSA.get(fila["root_cause_id"], GRIS))
        ws.cell(row=f, column=3, value=fila["dias"])
        ws.cell(row=f, column=4, value=fila["venta_perdida"])
        ws.cell(row=f, column=5, value=fila["pct"] / 100).number_format = "0.0%"
        ws.cell(row=f, column=6, value=fila["responsable"])
        for j in range(2, 7):
            ws.cell(row=f, column=j).border = BORDE
        f += 1

    f += 2
    _encabezado(ws, f, ["", "Responsable", "Días", "Venta perdida", "% del total", ""])
    f += 1
    for fila in resumen_por_responsable(en_alcance):
        ws.cell(row=f, column=2, value=fila["responsable"]).font = NEGRITA
        ws.cell(row=f, column=3, value=fila["dias"])
        ws.cell(row=f, column=4, value=fila["venta_perdida"])
        ws.cell(row=f, column=5, value=fila["pct"] / 100).number_format = "0.0%"
        for j in range(2, 7):
            ws.cell(row=f, column=j).border = BORDE
        f += 1

    # --- Por SKU-Tienda ---------------------------------------------------
    # Es la lista accionable: qué combinación está costando más y de quién es.
    ws = wb.create_sheet("Por SKU-Tienda")
    ws.sheet_view.showGridLines = False
    for j, w in enumerate([16, 9, 12, 12, 11, 15, 11, 15, 8, 26, 20, 60], 1):
        ws.column_dimensions[get_column_letter(j)].width = w

    ws["A1"] = "Desglose por SKU y tienda — ordenado por impacto"
    ws["A1"].font = TITULO
    ws["A2"] = ("La causa dominante es la que más venta perdida acumuló en el periodo. "
                "El desglose muestra cómo se repartió el impacto entre causas.")
    ws["A2"].font = CHICA

    # "OSA del periodo" reemplaza al viejo "OSA prom.", que promediaba sólo los
    # días con faltante y por eso salía 0.0 en TODOS los renglones (BOPS reporta
    # OSA binario: un día con faltante vale 0 por definición). Se agrega al lado
    # "días eval." para que el porcentaje se pueda leer contra su denominador.
    _encabezado(ws, 4, ["sku", "tienda", "días faltante", "días clasif.", "cobertura",
                        "venta perdida", "días eval.", "OSA del periodo", "RC",
                        "causa dominante", "responsable", "desglose de causas"])
    ws.row_dimensions[4].height = 32

    for i, dd in enumerate(diags):
        r = 5 + i
        desglose = " · ".join(
            f"{c}: ${v:,.0f}" for c, v in
            sorted(dd.desglose_causas.items(), key=lambda kv: kv[1], reverse=True))

        valores = [dd.sku, dd.tienda, dd.dias_con_faltante, dd.dias_clasificados,
                   dd.cobertura_pct / 100, dd.venta_perdida_total, dd.dias_evaluados,
                   None if dd.osa_periodo is None else dd.osa_periodo / 100,
                   dd.root_cause_id_dominante, dd.causa_dominante,
                   dd.responsable_dominante, desglose]
        for j, v in enumerate(valores, 1):
            c = ws.cell(row=r, column=j, value=v)
            c.border = BORDE
        ws.cell(row=r, column=5).number_format = "0.0%"
        ws.cell(row=r, column=6).number_format = '$#,##0.00'
        ws.cell(row=r, column=8).number_format = "0.0%"
        rc = ws.cell(row=r, column=9)
        rc.font = NEGRITA
        rc.fill = PatternFill("solid", fgColor=COLOR_CAUSA.get(dd.root_cause_id_dominante, GRIS))
        rc.alignment = Alignment(horizontal="center")
        if dd.cobertura_pct < 100:
            ws.cell(row=r, column=5).fill = PatternFill("solid", fgColor=AMBAR)

    ws.freeze_panes = "C5"
    ws.auto_filter.ref = f"A4:L{4 + len(diags)}"

    # --- Proveedor y citas ------------------------------------------------
    escribir_hoja_proveedor(wb, fu, en_alcance)

    # --- Cobertura y fuentes ---------------------------------------------
    ws = wb.create_sheet("Cobertura y fuentes")
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABCDEF", [4, 34, 14, 16, 16, 54]):
        ws.column_dimensions[col].width = w

    cob = cobertura_modelo(diagnosticos)
    cob["osa_general"] = osa_general(fu)
    cob["osa_alcance"] = osa_alcance(fu)

    ws["B2"] = "Cobertura del modelo"
    ws["B2"].font = TITULO
    f = 4
    if aviso:
        _banner(ws, f, 2, 6, aviso, alto=60)
        f += 2
    pct = lambda v: f"{v}%" if v is not None else "n/d"
    filas_resumen = [
        ("OSA del periodo (SKU del catálogo)", pct(cob["osa_alcance"])),
        ("Días con faltante que entregó BOPS", cob["casos_totales"]),
    ]
    if cob["casos_fuera_de_alcance"]:
        filas_resumen.insert(1, ("OSA sobre todo lo que entregó BOPS",
                                 pct(cob["osa_general"])))
    # Las dos coberturas sólo se separan cuando de verdad hay días fuera del
    # catálogo; si no, mostrar ambas sería ruido.
    if cob["casos_fuera_de_alcance"]:
        filas_resumen += [
            ("  de esos, fuera del catálogo de la tienda",
             f"{cob['casos_fuera_de_alcance']} "
             f"(${cob['venta_perdida_fuera_de_alcance']:,.2f})"),
            ("Días con faltante DENTRO del alcance", cob["casos_en_alcance"]),
            ("Días clasificados", cob["casos_clasificados"]),
            ("Cobertura sobre el alcance",
             f"{cob['cobertura_casos_alcance_pct']}%"),
            ("Cobertura sobre todo lo que entregó BOPS",
             f"{cob['cobertura_casos_pct']}%"),
            ("Venta perdida dentro del alcance",
             f"${cob['venta_perdida_en_alcance']:,.2f}"),
            ("Venta perdida clasificada", f"${cob['venta_perdida_clasificada']:,.2f}"),
            ("Cobertura sobre impacto (alcance)",
             f"{cob['cobertura_venta_perdida_alcance_pct']}%"),
        ]
    else:
        filas_resumen += [
            ("Días clasificados", cob["casos_clasificados"]),
            ("Cobertura sobre días", f"{cob['cobertura_casos_pct']}%"),
            ("Venta perdida total", f"${cob['venta_perdida_total']:,.2f}"),
            ("Venta perdida clasificada", f"${cob['venta_perdida_clasificada']:,.2f}"),
            ("Cobertura sobre impacto", f"{cob['cobertura_venta_perdida_pct']}%"),
        ]
    for k, v in filas_resumen:
        ws.cell(row=f, column=2, value=k).font = NEGRITA
        ws.cell(row=f, column=2).fill = PatternFill("solid", fgColor=GRIS)
        ws.cell(row=f, column=4, value=v)
        for j in range(2, 5):
            ws.cell(row=f, column=j).border = BORDE
        f += 1

    if cob["campos_que_bloquean"]:
        f += 2
        ws.cell(row=f, column=2, value="Qué dato bloqueó la clasificación").font = TITULO
        f += 2
        _encabezado(ws, f, ["", "Campo faltante", "Días", "Venta perdida sin explicar",
                            "", "A quién pedírselo"])
        f += 1
        vp_por_campo = cob["venta_perdida_por_campo_faltante"]
        de_quien = {
            "inventario_tienda": "Tableau — inventario en tienda",
            "transito_vigente": "CEDIS / Logística — transferencias",
            "envio_cedis_generado": "CEDIS / Logística — transferencias",
            "pedido_tienda_generado": "SIMA — pedidos de tienda",
            "via_resurtido": "Catálogo — vía de resurtido del SKU",
            "inventario_cedis": "CEDIS — inventario diario",
            "pedido_proveedor_generado": "Compras — pedidos a proveedor",
            "proveedor_cajas_pedidas": "Compras — pedidos a proveedor",
            "proveedor_cajas_entregadas": "Compras — citas de proveedor",
            "proveedor_cajas_confirmadas_cita": "Compras — citas de proveedor",
            "regla_matriz_para_cita_de_proveedor_no_vencida":
                "Decisión de negocio — la matriz no cubre el faltante previo a la cita",
            "regla_matriz_para_entrega_completa_con_cedis_en_cero":
                "Decisión de negocio — la matriz no cubre CEDIS en cero con entrega completa",
            FUERA_DE_CATALOGO:
                "BOPS — SKU de una división que este análisis no cubre. BOPS entrega todas las divisiones y el catálogo sólo trae Abarrotes. No es un dato faltante: son días fuera del alcance",
        }
        for campo, n in cob["campos_que_bloquean"].items():
            ws.cell(row=f, column=2, value=campo).font = Font(name="Consolas", size=10)
            ws.cell(row=f, column=3, value=n)
            ws.cell(row=f, column=4, value=vp_por_campo.get(campo, 0)).number_format = '$#,##0.00'
            ws.cell(row=f, column=6, value=de_quien.get(campo, "Revisar matriz")).alignment = WRAP
            for j in range(2, 7):
                ws.cell(row=f, column=j).border = BORDE
                ws.cell(row=f, column=j).fill = PatternFill("solid", fgColor=AMBAR)
            f += 1

    f += 2
    ws.cell(row=f, column=2, value="Fuentes leídas").font = TITULO
    f += 2
    _encabezado(ws, f, ["", "Hoja", "Filas", "Primera fecha", "Última fecha", "Equipo"])
    f += 1
    for hoja in HOJAS:
        n = fu.conteo.get(hoja, 0)
        d0, d1 = fu.rango.get(hoja, (None, None))
        ws.cell(row=f, column=2, value=hoja).font = Font(name="Consolas", size=10)
        ws.cell(row=f, column=3, value=n)
        ws.cell(row=f, column=4, value=d0.isoformat() if d0 else "")
        ws.cell(row=f, column=5, value=d1.isoformat() if d1 else "")
        ws.cell(row=f, column=6, value=HOJAS[hoja]["equipo"])
        for j in range(2, 7):
            ws.cell(row=f, column=j).border = BORDE
            if n == 0:
                ws.cell(row=f, column=j).fill = PatternFill("solid", fgColor=ROJO)
        f += 1

    if fu.advertencias:
        f += 2
        ws.cell(row=f, column=2, value="Advertencias de lectura").font = TITULO
        f += 2
        for a in fu.advertencias:
            c = ws.cell(row=f, column=2, value=a)
            c.alignment = WRAP
            c.fill = PatternFill("solid", fgColor=AMBAR)
            ws.merge_cells(start_row=f, start_column=2, end_row=f, end_column=6)
            f += 1

    wb.save(ruta)
    return cob, par


# ===========================================================================
# 4. CLI
# ===========================================================================

def main() -> int:
    ap = argparse.ArgumentParser(
        description="Lee el Excel de captura y clasifica el desabasto por causa raíz.")
    ap.add_argument("archivo",
                    help="Excel de captura llenado por los equipos.")
    ap.add_argument("csv", nargs="*", type=Path,
                    help="CSV de las hojas que ya no caben en el Excel "
                         "(TABLEAU_INV_TIENDA_*.csv, TABLEAU_VENTAS.csv).")
    ap.add_argument("-o", "--salida", default=None,
                    help="Excel de resultados. Por omisión, 'Resultado RCA - <archivo>'.")
    ap.add_argument("--umbral-osa", type=float, default=100.0,
                    help="Se analizan los días con OSA por debajo de este valor (default 100).")
    args = ap.parse_args()

    entrada = Path(args.archivo)
    if not entrada.exists():
        print(f"No se encontró el archivo: {entrada}", file=sys.stderr)
        return 1

    faltan = [c for c in args.csv if not c.exists()]
    if faltan:
        print(f"No se encontraron: {', '.join(str(c) for c in faltan)}", file=sys.stderr)
        return 1

    salida = Path(args.salida) if args.salida else \
        entrada.with_name(f"Resultado RCA - {entrada.stem}.xlsx")

    paquete = PaqueteFuentes.desde(entrada, args.csv)
    print(f"\nLeyendo  {entrada.name}")
    for hoja, rutas in paquete.csvs.items():
        print(f"         {hoja} <- {len(rutas)} CSV "
              f"({sum(r.stat().st_size for r in rutas) / 1e6:.0f} MB)")
    fu = leer_fuentes(paquete, args.umbral_osa)
    for hoja in HOJAS:
        n = fu.conteo.get(hoja, 0)
        d0, d1 = fu.rango.get(hoja, (None, None))
        rango = f"{d0} a {d1}" if d0 else "sin fechas"
        marca = "  " if n else "!!"
        print(f"  {marca} {hoja:<24} {n:>4} filas   {rango}")

    if fu.advertencias:
        print("\nAdvertencias:")
        for a in fu.advertencias:
            print(f"   - {a}")

    evidencias = derivar_evidencias(fu, args.umbral_osa)
    if not evidencias:
        osa_gral = osa_general(fu)
        if osa_gral is not None:
            print(f"\nOSA general del periodo: {osa_gral}%")
        print("\nNo hay días con faltante que analizar. Revisa la hoja BOPS_OSA.")
        return 1

    diagnosticos = clasificar(evidencias)   # una sola pasada del motor

    cob, par = escribir_resultado(salida, fu, evidencias, diagnosticos)

    if cob["osa_alcance"] is not None:
        print(f"\nOSA del periodo (SKU del catálogo): {cob['osa_alcance']}%")
        if cob["casos_fuera_de_alcance"] and cob["osa_general"] != cob["osa_alcance"]:
            print(f"OSA sobre todo lo que entregó BOPS: {cob['osa_general']}% "
                  f"— mezcla divisiones que el análisis no cubre")

    agua = waterfall_osa(fu, diagnosticos)
    if agua["escalones"]:
        print(f"\nDónde se pierde la disponibilidad  (100% -> {agua['osa_real']}%)")
        for e in agua["escalones"]:
            print(f"  -{e['puntos_osa']:>5.2f} pp   {e['causa']:<32} "
                  f"{e['dias']:>6,} días")

    print(f"\nBOPS entregó {cob['casos_totales']:,} días con faltante")
    if cob["casos_fuera_de_alcance"]:
        print(f"  {cob['casos_fuera_de_alcance']:,} fuera del catálogo de la tienda "
              f"(${cob['venta_perdida_fuera_de_alcance']:,.2f}) — no entran al análisis")
        print(f"  {cob['casos_en_alcance']:,} dentro del alcance · "
              f"clasificados {cob['casos_clasificados']:,} "
              f"({cob['cobertura_casos_alcance_pct']}%)")
        print(f"Venta perdida en alcance ${cob['venta_perdida_en_alcance']:,.2f} · "
              f"clasificada ${cob['venta_perdida_clasificada']:,.2f} "
              f"({cob['cobertura_venta_perdida_alcance_pct']}%)")
    else:
        print(f"  clasificados {cob['casos_clasificados']:,} "
              f"({cob['cobertura_casos_pct']}%)")
        print(f"Venta perdida ${cob['venta_perdida_total']:,.2f} · "
              f"clasificada ${cob['venta_perdida_clasificada']:,.2f} "
              f"({cob['cobertura_venta_perdida_pct']}%)")

    print("\nPareto por causa raíz")
    for causa, pct in par["pareto_por_causa"].items():
        print(f"  {pct:>5.1f}%   {causa}")

    print("\nPareto por responsable")
    for resp, pct in par["pareto_por_responsable"].items():
        print(f"  {pct:>5.1f}%   {resp}")

    if cob["campos_que_bloquean"]:
        print("\nDatos que bloquearon la clasificación")
        vp_campo = cob["venta_perdida_por_campo_faltante"]
        for campo, n in cob["campos_que_bloquean"].items():
            print(f"  {n:>3} días   ${vp_campo.get(campo, 0):>12,.2f}   falta {campo}")

    prov = desempeno_proveedores(fu)
    if prov and not fu.vacia("CITAS_PROV_CEDIS"):
        print("\nCumplimiento del proveedor en CEDIS")
        for d in prov:
            def p(v):
                return f"{v*100:5.1f}%" if v is not None else "   n/d"
            print(f"  {(d.nombre or d.proveedor_id)[:34]:<34} "
                  f"confirmado {p(d.pct_confirmado)}  "
                  f"cumplimiento {p(d.pct_cumplimiento)}  "
                  f"efectivo {p(d.pct_efectivo)}  "
                  f"({d.pedidos_sin_cita} pedidos sin cita, "
                  f"{d.citas_incumplidas} citas falladas)")

    print("\nTop SKU-Tienda por impacto")
    diags_st = diagnosticar_periodo(dentro_del_alcance(diagnosticos))
    for dd in diags_st[:10]:
        print(f"  ${dd.venta_perdida_total:>12,.2f}   SKU {dd.sku} / tienda {dd.tienda}   "
              f"[{dd.root_cause_id_dominante}] {dd.causa_dominante}")

    aviso = aviso_prioridad_3(fu)
    if aviso:
        print(f"\n{'!' * 78}")
        for linea in textwrap.wrap(aviso, 76):
            print(f"! {linea}")
        print("!" * 78)

    print(f"\nResultado escrito en: {salida.name}\n")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
