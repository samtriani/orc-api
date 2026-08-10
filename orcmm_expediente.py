"""Expediente de un SKU — todo lo que las fuentes dicen de él, en un solo lugar.

El pipeline entrega el veredicto de 30 mil días a la vez. Cuando uno de esos
días se ve raro —o cuando un SKU que debería aparecer no aparece— no hay forma
de auditarlo desde el Excel de salida: hay que ir a las ocho fuentes y cruzarlas
a mano. Esto hace ese cruce.

Contesta dos preguntas que el resultado no contesta:

  · Por qué el árbol dictaminó lo que dictaminó — con el dato crudo de cada
    hoja al lado del veredicto, no sólo la columna resumida.
  · Por qué un SKU no produjo ni una fila. Casi siempre es que BOPS_OSA nunca
    lo reportó, y eso no se ve en ningún lado: el SKU simplemente no está.

Uso:

    python orcmm_expediente.py 7506425626212 "layout.xlsx" datos/*.csv \\
        --tienda 287 --resultado "Resultado RCA.xlsx"

El layout es lo único obligatorio. Los CSV de Tableau y el Excel de resultado
son opcionales: sin ellos se omiten esas secciones y el resto sale igual.
"""
from __future__ import annotations

import argparse
import sys
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional

import openpyxl

from orcmm_fuentes_csv import agrupar_por_hoja, leer_csv

# Hoja del layout -> (índice de la columna con el SKU, para qué sirve la hoja).
# El índice cambia entre hojas porque no todas empiezan por SKU: en las de
# compras la primera columna es el folio.
HOJAS = [
    ("BOPS_OSA", 0, "OSA diario en anaquel — DE AQUÍ SALEN LOS DÍAS CON FALTANTE"),
    ("CEDIS_INVENTARIO", 0, "Inventario diario en CEDIS"),
    ("CEDIS_TRANSFERENCIAS", 0, "Envíos de CEDIS a tienda"),
    ("SIMA_PEDIDOS_TIENDA", 0, "Pedidos que la tienda generó"),
    ("COMPRAS_PEDIDOS_PROV", 1, "Pedidos a proveedor"),
    ("CITAS_PROV_CEDIS", 2, "Citas del proveedor en CEDIS"),
]

ANCHO = 78


def _iso(v) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return "" if v is None else str(v)


def _titulo(t: str) -> None:
    print(f"\n{'=' * ANCHO}\n{t}\n{'=' * ANCHO}")


def _tabla(encabezado, filas, tope: int = 60) -> None:
    """Imprime con separador de pipes. Se trunca a `tope` porque un SKU con año
    y medio de pedidos trae cientos de filas y lo que importa es el patrón."""
    cols = [i for i, h in enumerate(encabezado) if h]
    print("  " + " | ".join(str(encabezado[i]) for i in cols))
    print("  " + "-" * (ANCHO - 4))
    for r in filas[:tope]:
        print("  " + " | ".join(_iso(r[i]) if i < len(r) else "" for i in cols))
    if len(filas) > tope:
        print(f"  … y {len(filas) - tope:,} filas más (total {len(filas):,})")
    else:
        print(f"  total: {len(filas):,} filas")


def _seccion_catalogo(wb, sku: str, tienda: Optional[str]) -> None:
    """El catálogo define la vía de resurtido y el CEDIS surtidor, que son los
    que deciden por qué rama del árbol baja el SKU. Si no está aquí, el día
    queda fuera del alcance sin importar lo que digan las demás hojas."""
    _titulo("CATÁLOGO — define vía, CEDIS surtidor y vigencia")
    encabezado, filas = None, []
    for i, r in enumerate(wb["CATALOGO"].iter_rows(min_row=3, values_only=True)):
        if i == 0:
            encabezado = r
            continue
        if r[0] is None or str(r[0]) != sku:
            continue
        if tienda and str(r[2]) != tienda:
            continue
        filas.append(r)

    if not filas:
        print("  (no aparece — cualquier día de este SKU sale como")
        print("   'sku_fuera_del_catalogo_de_la_tienda' y no se clasifica)")
        return
    for r in filas:
        print("  " + " · ".join(f"{h}={_iso(v)}" for h, v in zip(encabezado, r)
                                if h and v not in (None, "")))


def _seccion_hoja(wb, hoja: str, col_sku: int, desc: str,
                  sku: str, tienda: Optional[str]) -> int:
    _titulo(f"{hoja} — {desc}")
    if hoja not in wb.sheetnames:
        print("  (la hoja no viene en este layout)")
        return 0

    encabezado, filas = None, []
    for i, r in enumerate(wb[hoja].iter_rows(min_row=3, values_only=True)):
        if i == 0:
            encabezado = r
            continue
        if col_sku >= len(r) or r[col_sku] is None or str(r[col_sku]) != sku:
            continue
        # Sólo las hojas de tienda tienen columna de tienda en la posición 1;
        # en las de compras esa posición es el proveedor, así que no se filtra.
        if tienda and hoja in ("BOPS_OSA", "SIMA_PEDIDOS_TIENDA") \
                and str(r[1]) != tienda:
            continue
        filas.append(r)

    if not filas:
        print("  (sin filas para este SKU)")
        return 0
    _tabla(encabezado, filas)
    return len(filas)


def _seccion_csv(rutas: List[Path], hoja: str, titulo: str, campo: str,
                 sku: str, tienda: Optional[str]) -> None:
    _titulo(titulo)
    if not rutas:
        print("  (no se pasaron CSV de esta hoja)")
        return

    filas = [f for f in leer_csv(rutas, hoja)
             if str(f.get("sku")) == sku
             and (not tienda or str(f.get("tienda")) == tienda)]
    if not filas:
        print("  (sin filas para este SKU)")
        return

    filas.sort(key=lambda f: (f.get("fecha") or date.min, str(f.get("tienda"))))
    vals = [f.get(campo) for f in filas if f.get(campo) is not None]
    ceros = sum(1 for v in vals if float(v) == 0)
    if vals:
        print(f"  {len(filas):,} días · {ceros} en cero · "
              f"mínimo {min(float(v) for v in vals):g} · "
              f"máximo {max(float(v) for v in vals):g}")
    for f in filas:
        v = f.get(campo)
        # El cero es el dato que importa: es el día en que no había producto.
        marca = "   <-- EN CERO" if v is not None and float(v) == 0 else ""
        print(f"  {_iso(f.get('fecha'))}  tienda {f.get('tienda')}  "
              f"{'' if v is None else f'{float(v):>10.2f}'}{marca}")


def _seccion_veredicto(ruta: Optional[Path], sku: str, tienda: Optional[str],
                       hubo_bops: bool) -> None:
    _titulo("VEREDICTO DEL RCA — hoja 'Clasificacion diaria' del resultado")
    if ruta is None:
        print("  (no se pasó --resultado)")
        return
    if not ruta.exists():
        print(f"  (no existe {ruta.name} — correr antes orcmm_pipeline.py)")
        return

    wb = openpyxl.load_workbook(ruta, read_only=True)
    if "Clasificacion diaria" not in wb.sheetnames:
        print("  (el archivo no trae la hoja 'Clasificacion diaria')")
        return

    encabezado, filas = None, []
    for r in wb["Clasificacion diaria"].iter_rows(values_only=True):
        if encabezado is None:
            if r and r[0] == "sku":
                encabezado = r
            continue
        if r[0] is not None and str(r[0]) == sku:
            if tienda and str(r[1]) != tienda:
                continue
            filas.append(r)

    if not filas:
        print("  (ninguna fila — el SKU no produjo días clasificables)")
        # La causa casi siempre es ésta, y sin decirlo se lee como un error.
        if not hubo_bops:
            print()
            print("  BOPS_OSA no trae una sola fila de este SKU, así que no hay")
            print("  ningún día con faltante que clasificar. No es una falla del")
            print("  modelo: el análisis arranca de los días que BOPS reporta.")
        return
    _tabla(encabezado, filas, tope=80)


def main() -> int:
    ap = argparse.ArgumentParser(
        description="Expediente de un SKU a partir de todas las fuentes.")
    ap.add_argument("sku")
    ap.add_argument("archivo", type=Path, help="El layout .xlsx")
    ap.add_argument("csv", nargs="*", type=Path,
                    help="Los CSV de Tableau, si se tienen.")
    ap.add_argument("--tienda", default=None,
                    help="Acota a una tienda. Sin esto salen todas.")
    ap.add_argument("--resultado", type=Path, default=None,
                    help="Excel de salida del pipeline, para ver el veredicto.")
    a = ap.parse_args()

    sku = a.sku.strip()
    if not a.archivo.exists():
        print(f"No existe el layout: {a.archivo}", file=sys.stderr)
        return 2

    print(f"EXPEDIENTE DEL SKU {sku}"
          + (f"  ·  tienda {a.tienda}" if a.tienda else "  ·  todas las tiendas"))
    print(f"Layout: {a.archivo.name}")

    wb = openpyxl.load_workbook(a.archivo, read_only=True, data_only=True)
    _seccion_catalogo(wb, sku, a.tienda)

    hubo_bops = False
    for hoja, col, desc in HOJAS:
        n = _seccion_hoja(wb, hoja, col, desc, sku, a.tienda)
        if hoja == "BOPS_OSA":
            hubo_bops = n > 0

    por_hoja, _sueltos = agrupar_por_hoja(a.csv)
    _seccion_csv(por_hoja.get("TABLEAU_INV_TIENDA", []), "TABLEAU_INV_TIENDA",
                 "TABLEAU_INV_TIENDA — existencia diaria en tienda",
                 "existencia_piezas", sku, a.tienda)
    _seccion_csv(por_hoja.get("TABLEAU_VENTAS", []), "TABLEAU_VENTAS",
                 "TABLEAU_VENTAS — venta diaria", "importe_venta", sku, a.tienda)

    _seccion_veredicto(a.resultado, sku, a.tienda, hubo_bops)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
