"""ORCMM — carga las fuentes de captura a Postgres (Neon).

No re-sube el Excel en cada corrida: esto guarda las 9 hojas del layout en
Postgres de forma permanente, por UPSERT sobre la llave natural de cada
una (idempotente — seguro de volver a correr con la misma entrega o con
una reentrega corregida).

Reutiliza el mismo parseo que ya usa el pipeline de clasificación
(orcmm_pipeline.leer_hoja, orcmm_fuentes_csv.leer_csv) — a diferencia de
ese pipeline, aquí se lee CADA fila (llaves=None), no sólo los días con
faltante: la base es almacenamiento crudo permanente, no una optimización
de memoria para una corrida puntual.

Uso:
    python orcmm_etl_carga.py layout.xlsx TABLEAU_INV_TIENDA_1.csv ... TABLEAU_VENTAS.csv
    python orcmm_etl_carga.py layout.xlsx --forzar       # cargar aunque haya errores de validación
    python orcmm_etl_carga.py layout.xlsx --sin-validar  # saltar validar_archivo por completo
"""
from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

from orcmm_db import (TABLAS, columnas_de, conectar, registrar_carga_fin,
                       registrar_carga_inicio, upsert_lote)
from orcmm_fuentes_csv import leer_csv
from orcmm_layout_spec import HOJAS
from orcmm_pipeline import PaqueteFuentes, leer_hoja
from orcmm_validar_layout import validar_archivo

TAMANO_LOTE = 5000


def preparar_fila(hoja: str, fila: dict) -> dict:
    """Aplica el conversor de tipo de cada campo. Idempotente: si la fila ya
    venía tipada (como las de leer_csv), no cambia nada."""
    return {col: conv(fila.get(campo)) for campo, col, conv in columnas_de(hoja)}


def cargar_hoja_excel(cur, wb, hoja: str, adv: list) -> int:
    filas = leer_hoja(wb, hoja, adv)
    tabla, llave = TABLAS[hoja]["tabla"], TABLAS[hoja]["llave"]
    columnas_bd = [c for _, c, _ in columnas_de(hoja)]
    total = 0
    for i in range(0, len(filas), TAMANO_LOTE):
        lote = [preparar_fila(hoja, f) for f in filas[i:i + TAMANO_LOTE]]
        total += upsert_lote(cur, tabla, columnas_bd, llave, lote, adv)
    return total


def cargar_hoja_csv(cur, rutas, hoja: str, adv: list) -> int:
    tabla, llave = TABLAS[hoja]["tabla"], TABLAS[hoja]["llave"]
    columnas_bd = [c for _, c, _ in columnas_de(hoja)]
    total, buffer = 0, []
    for f in leer_csv(rutas, hoja, llaves=None):   # None: se guarda TODO, no sólo lo filtrable
        buffer.append(preparar_fila(hoja, f))
        if len(buffer) >= TAMANO_LOTE:
            total += upsert_lote(cur, tabla, columnas_bd, llave, buffer, adv)
            buffer.clear()
    if buffer:
        total += upsert_lote(cur, tabla, columnas_bd, llave, buffer, adv)
    return total


def cargar(paquete: PaqueteFuentes, dsn: str | None, sin_validar: bool, forzar: bool) -> int:
    adv: list[str] = []
    validado = False
    if not sin_validar:
        rep = validar_archivo(paquete.xlsx, [p for v in paquete.csvs.values() for p in v])
        validado = True
        if rep.errores and not forzar:
            print("Validación con errores — usa --forzar para cargar de todas formas:", file=sys.stderr)
            for e in rep.errores:
                print(f"  - {e}", file=sys.stderr)
            return 1
        adv.extend(rep.advertencias)

    wb = openpyxl.load_workbook(paquete.xlsx, data_only=True, read_only=True)
    conn = conectar(dsn)
    t_inicio = time.time()
    try:
        with conn, conn.cursor() as cur:
            carga_id = registrar_carga_inicio(
                cur, paquete.xlsx.name, [p.name for v in paquete.csvs.values() for p in v])

        estado, filas_por_hoja = "ok", {}
        for hoja in HOJAS:
            t0 = time.time()
            try:
                with conn, conn.cursor() as cur:   # una transacción por hoja
                    rutas = paquete.csvs.get(hoja)
                    # Mismo criterio que leer_fuentes: si la hoja llegó como CSV
                    # y también como pestaña del Excel, gana el Excel.
                    if rutas and hoja in wb.sheetnames:
                        adv.append(f"{hoja}: llegó como CSV y también como hoja del Excel. Se usa el Excel.")
                        rutas = None
                    if rutas:
                        n = cargar_hoja_csv(cur, rutas, hoja, adv)
                    elif hoja in wb.sheetnames:
                        n = cargar_hoja_excel(cur, wb, hoja, adv)
                    else:
                        n = 0
                    filas_por_hoja[hoja] = n
                print(f"  ok  {hoja:<24} {n:>9,} filas  ({time.time() - t0:.1f}s)")
            except Exception as e:
                estado = "parcial"
                adv.append(f"{hoja}: la carga falló y se saltó — {e}")
                print(f"  !!  {hoja:<24} FALLÓ: {e}", file=sys.stderr)
                # una hoja rota no bloquea las demás

        with conn, conn.cursor() as cur:
            registrar_carga_fin(cur, carga_id, estado, filas_por_hoja, adv,
                                 validado, forzar, time.time() - t_inicio)
        return 0 if estado == "ok" else 1
    finally:
        conn.close()


def main() -> int:
    ap = argparse.ArgumentParser(description="Carga las fuentes de ORCMM a Postgres (Neon).")
    ap.add_argument("xlsx", type=Path, help="Excel del layout de captura.")
    ap.add_argument("csv", nargs="*", type=Path, help="CSV sueltos (TABLEAU_INV_TIENDA_*.csv, etc).")
    ap.add_argument("--sin-validar", action="store_true", help="Saltar validar_archivo por completo.")
    ap.add_argument("--forzar", action="store_true", help="Cargar aunque la validación reporte errores.")
    ap.add_argument("--dsn", default=None, help="Por omisión, DATABASE_URL del entorno / .env.")
    args = ap.parse_args()

    paquete = PaqueteFuentes.desde(args.xlsx, args.csv)
    return cargar(paquete, args.dsn, args.sin_validar, args.forzar)


if __name__ == "__main__":
    load_dotenv()
    raise SystemExit(main())
