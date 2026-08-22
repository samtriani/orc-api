"""Escritura de Excel en streaming, con la misma API que ya usa el pipeline.

POR QUÉ EXISTE. openpyxl arma el libro entero en memoria antes de guardarlo:
cada celda es un objeto de Python. La hoja "Clasificación diaria" trae 189,470
renglones por 24 columnas —4.5 millones de celdas— y eso costaba **263
segundos y 3 GB de RAM** medidos sobre Coyoacán marzo. El archivo resultante
pesa 16 MB: lo caro no es el archivo, es construirlo objeto por objeto.

QUÉ HACE. Envuelve xlsxwriter en modo `constant_memory`, que escribe renglón
por renglón a disco y no guarda más que el renglón en curso.

POR QUÉ ADAPTADOR Y NO REESCRITURA. El armado de las cinco hojas son 483
líneas de posiciones, colores y anchos, y es un entregable de cliente: portarlo
a mano a otra librería es justo donde se cuelan las regresiones que nadie ve
hasta que el cliente abre el archivo. Aquí se cambia el MOTOR y se deja la
lógica intacta — el pipeline sigue escribiendo `ws.cell(...).font = NEGRITA`
como siempre.

LA REGLA DEL MODO STREAMING: los renglones de una hoja hay que escribirlos en
orden. Las hojas chicas se guardan enteras en memoria y se vuelcan ordenadas
al cerrar, así el orden deja de importar para ellas; sólo la diaria escribe
al vuelo, que es la única donde el volumen lo exige.
"""
from __future__ import annotations

from typing import Dict, Optional

import xlsxwriter
from openpyxl.utils import column_index_from_string, get_column_letter


def _rgb(color) -> Optional[str]:
    """El color de un objeto de openpyxl, en el '#RRGGBB' que espera xlsxwriter.

    openpyxl los guarda como 'FFC7CE' o 'FFFFC7CE' (con alfa por delante). Se
    tira el alfa: xlsxwriter no lo usa y dejarlo pinta el color equivocado.
    """
    if color is None:
        return None
    valor = getattr(color, "rgb", color)
    if not isinstance(valor, str):
        return None
    valor = valor.lstrip("#")
    if len(valor) == 8:
        valor = valor[2:]
    return f"#{valor}" if len(valor) == 6 else None


class _Celda:
    """Una celda con sus atributos, antes de escribirse.

    Existe porque el pipeline hace `c = ws.cell(...)` y DESPUÉS le pone el
    formato. La celda no se puede escribir en el momento de crearla: hay que
    esperar a que el renglón se cierre.
    """

    __slots__ = ("valor", "font", "fill", "alignment", "border", "number_format")

    def __init__(self, valor=None):
        self.valor = valor
        self.font = None
        self.fill = None
        self.alignment = None
        self.border = None
        self.number_format = None

    # El pipeline lee `.value` en un par de lugares.
    @property
    def value(self):
        return self.valor

    @value.setter
    def value(self, v):
        self.valor = v

    def clave(self) -> tuple:
        """Firma del formato, para no crear uno nuevo por cada celda.

        Sin esto xlsxwriter guardaría millones de formatos idénticos y
        volveríamos al problema que este módulo viene a resolver.
        """
        f, a = self.font, self.alignment
        return (
            bool(f and f.bold), (f.size if f else None), bool(f and f.italic),
            (f.name if f else None), _rgb(f.color) if f else None,
            _rgb(self.fill.fgColor) if self.fill is not None else None,
            (a.horizontal if a else None), (a.vertical if a else None),
            bool(a and a.wrap_text),
            self.border is not None,
            self.number_format,
        )


class _Anchos:
    """`ws.column_dimensions['C'].width = 12`, tal cual lo escribe el pipeline."""

    def __init__(self, hoja: "Hoja"):
        self._hoja = hoja

    def __getitem__(self, letra):
        return _AnchoCol(self._hoja, letra)


class _AnchoCol:
    def __init__(self, hoja, letra):
        self._hoja, self._letra = hoja, letra

    @property
    def width(self):
        return None

    @width.setter
    def width(self, w):
        i = column_index_from_string(self._letra) - 1
        self._hoja._ws.set_column(i, i, w)


class _Altos:
    def __init__(self, hoja: "Hoja"):
        self._hoja = hoja

    def __getitem__(self, fila):
        return _AltoFila(self._hoja, fila)


class _AltoFila:
    def __init__(self, hoja, fila):
        self._hoja, self._fila = hoja, fila

    @property
    def height(self):
        return None

    @height.setter
    def height(self, h):
        self._hoja._altos[self._fila] = h


class _Filtro:
    def __init__(self, hoja: "Hoja"):
        self._hoja = hoja

    @property
    def ref(self):
        return None

    @ref.setter
    def ref(self, r):
        self._hoja._filtro = r


class _Vista:
    def __init__(self):
        self.showGridLines = True


class Hoja:
    """Una hoja con la parte de la API de openpyxl que el pipeline usa."""

    def __init__(self, libro: "Libro", ws, streaming: bool):
        self._libro = libro
        self._ws = ws
        self._streaming = streaming
        self._filas: Dict[int, Dict[int, _Celda]] = {}
        self._altos: Dict[int, float] = {}
        self._merges: Dict[int, tuple] = {}
        self._filtro: Optional[str] = None
        self._congelar: Optional[str] = None
        self._ultima_volcada = 0

        self.column_dimensions = _Anchos(self)
        self.row_dimensions = _Altos(self)
        self.auto_filter = _Filtro(self)
        self.sheet_view = _Vista()

    # -- API que consume el pipeline ---------------------------------------

    def cell(self, row: int, column: int, value=None) -> _Celda:
        fila = self._filas.get(row)
        if fila is None:
            # En streaming, pasar a un renglón nuevo cierra los anteriores.
            # Es lo que permite no guardar la hoja entera.
            if self._streaming and row > self._ultima_volcada:
                self._volcar_hasta(row - 1)
            fila = self._filas[row] = {}
        c = fila.get(column)
        if c is None:
            c = fila[column] = _Celda(value)
        elif value is not None:
            c.valor = value
        return c

    def __setitem__(self, ref: str, valor):
        col = "".join(ch for ch in ref if ch.isalpha())
        fil = "".join(ch for ch in ref if ch.isdigit())
        self.cell(row=int(fil), column=column_index_from_string(col), value=valor)

    def __getitem__(self, ref: str) -> _Celda:
        col = "".join(ch for ch in ref if ch.isalpha())
        fil = "".join(ch for ch in ref if ch.isdigit())
        return self.cell(row=int(fil), column=column_index_from_string(col))

    def merge_cells(self, start_row, start_column, end_row, end_column):
        # Por renglón: el merge se aplica cuando ese renglón se vuelca, no al
        # cerrar. En streaming el renglón ya se fue y escribirlo después
        # rompería el orden creciente que exige constant_memory.
        self._merges[start_row] = (start_column, end_row, end_column)

    @property
    def freeze_panes(self):
        return self._congelar

    @freeze_panes.setter
    def freeze_panes(self, ref):
        self._congelar = ref

    # -- volcado ------------------------------------------------------------

    def _volcar_hasta(self, hasta: int) -> None:
        for r in sorted(f for f in self._filas if f <= hasta):
            self._escribir_fila(r, self._filas.pop(r))
            self._ultima_volcada = max(self._ultima_volcada, r)

    def _escribir_fila(self, r: int, celdas: Dict[int, _Celda]) -> None:
        alto = self._altos.get(r)
        if alto is not None:
            self._ws.set_row(r - 1, alto)

        merge = self._merges.pop(r, None)
        if merge is not None:
            c1, r2, c2 = merge
            c = celdas.pop(c1, None)
            # merge_range escribe la celda de la esquina él mismo, así que se
            # saca de la lista para no escribirla dos veces.
            self._ws.merge_range(r - 1, c1 - 1, r2 - 1, c2 - 1,
                                 c.valor if c else None,
                                 self._libro.formato(c) if c else None)

        vacia = _Celda().clave()
        for col, c in celdas.items():
            if c.valor is None and c.clave() == vacia:
                continue                      # celda vacía y sin formato
            self._ws.write(r - 1, col - 1, c.valor, self._libro.formato(c))

    def cerrar(self) -> None:
        # Un merge en un renglón sin celdas no se volcaría solo: se le da una
        # fila vacía para que pase por _escribir_fila.
        for r in self._merges:
            self._filas.setdefault(r, {})
        self._volcar_hasta(max(self._filas) if self._filas else 0)

        if self._congelar:
            col = "".join(ch for ch in self._congelar if ch.isalpha())
            fil = "".join(ch for ch in self._congelar if ch.isdigit())
            self._ws.freeze_panes(int(fil) - 1, column_index_from_string(col) - 1)
        if self._filtro:
            self._ws.autofilter(self._filtro)
        if not self.sheet_view.showGridLines:
            self._ws.hide_gridlines(2)


class Libro:
    """Un libro de Excel escrito en streaming.

    `hoja(titulo, streaming=True)` es la que no cabe en memoria; las demás se
    guardan y se vuelcan al cerrar, para que el orden de escritura no importe.
    """

    def __init__(self, ruta):
        # constant_memory: cada renglón se escribe a disco y se olvida. Es
        # todo el punto del módulo.
        self._wb = xlsxwriter.Workbook(str(ruta), {"constant_memory": True})
        self._formatos: Dict[tuple, object] = {}
        self._hojas: list = []

    def hoja(self, titulo: str, streaming: bool = False) -> Hoja:
        h = Hoja(self, self._wb.add_worksheet(titulo[:31]), streaming)
        self._hojas.append(h)
        return h

    def formato(self, c: _Celda):
        """Un formato de xlsxwriter por combinación distinta, no por celda."""
        k = c.clave()
        f = self._formatos.get(k)
        if f is not None:
            return f

        props: dict = {}
        if c.font is not None:
            if c.font.bold:
                props["bold"] = True
            if c.font.italic:
                props["italic"] = True
            if c.font.size:
                props["size"] = float(c.font.size)
            if c.font.name:
                props["font_name"] = c.font.name
            color = _rgb(c.font.color)
            if color:
                props["font_color"] = color
        if c.fill is not None:
            relleno = _rgb(c.fill.fgColor)
            if relleno:
                props["bg_color"] = relleno
                props["pattern"] = 1
        if c.alignment is not None:
            if c.alignment.horizontal:
                props["align"] = c.alignment.horizontal
            if c.alignment.vertical:
                props["valign"] = {"center": "vcenter", "top": "top",
                                   "bottom": "bottom"}.get(c.alignment.vertical,
                                                           c.alignment.vertical)
            if c.alignment.wrap_text:
                props["text_wrap"] = True
        if c.border is not None:
            props["border"] = 1
            props["border_color"] = "#BFBFBF"
        if c.number_format:
            props["num_format"] = c.number_format

        f = self._formatos[k] = self._wb.add_format(props) if props else None
        return f

    def cerrar(self) -> None:
        for h in self._hojas:
            h.cerrar()
        self._wb.close()
