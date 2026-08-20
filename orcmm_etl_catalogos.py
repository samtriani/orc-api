"""ORCMM — carga los catálogos informativos (sucursales + SKU por tienda) a
Postgres. Son tablas de referencia/reporte: el motor de clasificación
sigue usando `catalogo` (la hoja del layout, con cedis_surtidor), no éstas.

No usan orcmm_layout_spec.HOJAS porque no son parte del layout de captura —
son archivos aparte, con su propio formato (encabezados en la fila 1, un
archivo por tienda para el catálogo de SKU).

Uso:
    python orcmm_etl_catalogos.py --sucursales "Listado_sucursales....xlsx"
    python orcmm_etl_catalogos.py --sku Catalogo_Abarrotes_Coyoacan.xlsx Catalogo_Abarrotes_Centenario.xlsx ...
    python orcmm_etl_catalogos.py --sucursales sucursales.xlsx --sku *.xlsx
"""
from __future__ import annotations

import argparse
import csv
import re
import sys
import time
import unicodedata
from pathlib import Path
from typing import Optional

import openpyxl
from dotenv import load_dotenv

from orcmm_db import (conectar, registrar_carga_fin, registrar_carga_inicio,
                       upsert_lote)
from orcmm_fuentes_csv import detectar_delimitador, detectar_encoding
from orcmm_pipeline import _decimal, _entero, _fecha, _texto

TAMANO_LOTE = 5000


def _es_na(v) -> bool:
    """Estos catálogos (Centenario, La Herradura, Plaza Carso, SMA) usan el
    texto literal 'NA' para dato ausente en columnas numéricas — distinto
    del '#N/A' de fórmula de Excel que ya reconoce _es_vacio en
    orcmm_pipeline. Se filtra aquí, sin tocar el conversor compartido."""
    return isinstance(v, str) and v.strip().upper() == "NA"


def _entero_na(v) -> Optional[int]:
    return None if _es_na(v) else _entero(v)


def _codigo(v) -> Optional[str]:
    """sku/tienda: en estos archivos llegan como número de Excel (9010,
    287.0, ...). Se normaliza a texto sin '.0' cuando es un entero limpio;
    si no lo es (código alfanumérico), se guarda tal cual."""
    if _es_na(v):
        return None
    n = _entero(v)
    return str(n) if n is not None else _texto(v)


# Encabezado real -> columna bd. Los 5 archivos de catálogo de SKU NO traen
# el mismo encabezado entre sí (verificado contra los 5 reales): Coyoacán
# trae "Codigo" Y "Código de Barras" (idénticos siempre, confirmado), otros
# traen sólo uno de los dos o una columna "code" al frente, y 3 de los 5 no
# traen "Línea"/"Vía" en absoluto. Por eso se empareja por NOMBRE con alias
# (mismo criterio que orcmm_layout_spec.ALIAS_ENCABEZADOS), no por posición
# fija — sólo sku/tienda son obligatorios, el resto que falte queda NULL.
def _normalizar(s) -> str:
    """Encabezado comparable: sin acentos, sin mayúsculas, sin espacios de más.

    Los acentos SE QUITAN a propósito. Antes se comparaba letra por letra y
    "Código" no calzaba con la llave "codigo": la columna se ignoraba en
    silencio y, por ser obligatoria, el archivo entero se abortaba con un
    "faltan columnas". Las llaves de los ALIAS pasan por aquí también, así que
    da igual con qué acentuación venga el archivo.
    """
    limpio = " ".join(str(s or "").split()).strip().lower()
    return "".join(c for c in unicodedata.normalize("NFD", limpio)
                   if unicodedata.category(c) != "Mn")


def _con_llaves_limpias(d: dict) -> dict:
    """Pasa las llaves de un mapa de alias por la misma regla que el
    encabezado del archivo, para que los dos lados se comparen igual."""
    return {_normalizar(k): v for k, v in d.items()}


def _sin_prefijo(v) -> Optional[str]:
    """Quita el prefijo numérico de la jerarquía comercial.

    Las entregas viejas traían "1 - ABARROTES" y "13 - VINOS Y LICORES (MAS
    DE 20 GL)"; la nueva trae el texto limpio. Sin normalizar, el MISMO valor
    aparecería dos veces en el filtro —una con número y otra sin— que es peor
    que cualquiera de las dos formas.

    El patrón exige dígitos, espacios, guion y espacios al inicio, para no
    morder un nombre que legítimamente empiece con número ("7 UP SABOR...").
    """
    t = _texto(v)
    return re.sub(r"^\s*\d+\s*-\s*", "", t) if t else t


ALIAS_SUCURSALES = _con_llaves_limpias({
    "formato": "formato",
    "no. tienda": "tienda",
    "nombre de la tienda": "nombre",
    "dirección": "direccion",
    "cp": "cp",
})
CONVERTIDORES_SUCURSALES = {
    "formato": _texto, "tienda": _codigo, "nombre": _texto,
    "direccion": _texto, "cp": _texto,
}
REQUERIDOS_SUCURSALES = {"tienda"}
COLUMNAS_BD_SUCURSALES = ["tienda", "formato", "nombre", "direccion", "cp"]

ALIAS_SKU_TIENDA = _con_llaves_limpias({
    "fecha inicial": "fecha_inicial",
    "tienda no": "tienda",
    "numero tienda": "tienda",
    "tienda nombre": "tienda_nombre",
    "nombre tienda": "tienda_nombre",
    "codigo": "sku",
    "code": "sku",
    "código de barras": "sku",
    "artículo nombre": "articulo_nombre",
    "división": "division",
    "descripción": "articulo_nombre",
    "grupo sección": "grupo_seccion",
    "sección": "grupo_seccion",
    "categoría": "categoria",
    "subcategoría": "subcategoria",
    "id proveedor": "proveedor_id",
    "proveedor": "proveedor_nombre",
    "proveedor nombre": "proveedor_nombre",
    "marca": "marca",
    "resurtido tipo": "resurtido_tipo",
    "resurtido frec": "resurtido_frec",
    "unidades de empaque": "unidades_empaque",
    "resurtido": "resurtido",
    "catálogo": "catalogo_activo",
    "linea o i&o": "linea_io",
    "via 1 o via 2": "via_resurtido",
    "vía": "via_resurtido",
})
CONVERTIDORES_SKU_TIENDA = {
    "fecha_inicial": _fecha, "tienda": _codigo, "tienda_nombre": _texto, "sku": _codigo,
    "articulo_nombre": _texto, "division": _sin_prefijo, "grupo_seccion": _sin_prefijo,
    "categoria": _sin_prefijo, "subcategoria": _sin_prefijo,
    "proveedor_id": _codigo, "marca": _texto,
    "proveedor_nombre": _texto, "resurtido_tipo": _texto, "resurtido_frec": _texto,
    "unidades_empaque": _entero_na, "resurtido": _entero_na, "catalogo_activo": _entero_na,
    "linea_io": _texto, "via_resurtido": _texto,
}
REQUERIDOS_SKU_TIENDA = {"sku", "tienda"}
COLUMNAS_BD_SKU_TIENDA = ["sku", "tienda", "tienda_nombre", "articulo_nombre", "division",
                          "grupo_seccion", "categoria", "subcategoria",
                          "proveedor_id", "proveedor_nombre", "marca",
                          "resurtido_tipo", "resurtido_frec",
                          "unidades_empaque", "resurtido", "catalogo_activo", "linea_io",
                          "via_resurtido", "fecha_inicial"]


def _filas_crudas(ruta: Path):
    """(encabezado, filas) de un .xlsx o un .csv.

    El catálogo comercial llega como CSV y el resto de los catálogos como
    Excel, así que se resuelve por extensión en vez de pedir que todos
    vengan en el mismo formato.

    Para el CSV se reutiliza la detección de `orcmm_fuentes_csv`: los
    archivos de La Comer salen en UTF-16 con BOM y a veces separados por
    TAB, así que asumir "UTF-8 con comas" falla en silencio — el archivo se
    lee como una sola columna de basura y las obligatorias "faltan".
    """
    if ruta.suffix.lower() != ".csv":
        wb = openpyxl.load_workbook(ruta, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        filas = ws.iter_rows(min_row=1, values_only=True)
        return next(filas), filas

    enc = detectar_encoding(ruta)
    with ruta.open("r", encoding=enc, newline="", errors="replace") as f:
        primera = f.readline()
        delim = detectar_delimitador(primera)
        encabezado = next(csv.reader([primera], delimiter=delim), [])
        cuerpo = list(csv.reader(f, delimiter=delim))
    return encabezado, cuerpo


def _leer_hoja_por_alias(ruta: Path, alias: dict, convertidores: dict,
                          requeridos: set, adv: list) -> list[dict]:
    """Encabezados en la fila 1, datos desde la fila 2 (a diferencia del
    layout de captura, que usa fila 3 / fila 6). Empareja cada columna por
    nombre normalizado contra `alias`; si dos encabezados distintos apuntan
    a la misma columna bd (p. ej. Codigo y Código de Barras -> sku), el que
    aparece más a la derecha en el archivo pisa al anterior."""
    encabezado, cuerpo = _filas_crudas(ruta)

    columnas = [(i, alias[_normalizar(h)]) for i, h in enumerate(encabezado)
                if _normalizar(h) in alias]
    presentes = {c for _, c in columnas}
    faltan = requeridos - presentes
    if faltan:
        adv.append(f"{ruta.name}: faltan columnas obligatorias {sorted(faltan)} "
                    f"(encabezado real: {[_normalizar(h) for h in encabezado if h]}). "
                    f"Se aborta este archivo.")
        return []

    filas = []
    for fila in cuerpo:
        if all(v is None or v == "" for v in fila):
            continue
        d = {col: convertidores[col](fila[i]) for i, col in columnas}
        filas.append(d)
    return filas


def cargar_sucursales(cur, ruta: Path, adv: list) -> int:
    filas = _leer_hoja_por_alias(ruta, ALIAS_SUCURSALES, CONVERTIDORES_SUCURSALES,
                                  REQUERIDOS_SUCURSALES, adv)
    total = 0
    for i in range(0, len(filas), TAMANO_LOTE):
        total += upsert_lote(cur, "sucursales", COLUMNAS_BD_SUCURSALES, ["tienda"],
                              filas[i:i + TAMANO_LOTE], adv)
    return total


def cargar_catalogo_sku_tienda(cur, ruta: Path, adv: list) -> int:
    """Carga o COMPLEMENTA el catálogo comercial.

    Sólo se escriben las columnas que el archivo trae de verdad. Es la
    diferencia entre complementar y pisar: una entrega puede traer la
    jerarquía comercial (división, categoría, marca) sin traer lo operativo
    (via_resurtido, unidades_empaque, catálogo activo), y al revés.

    Sin esto, `upsert_lote` mandaba None en cada columna ausente y el SET las
    escribía encima: cargar el catálogo comercial habría BORRADO
    via_resurtido y compañía en las 107,339 filas ya cargadas, sin un solo
    error que lo delatara.
    """
    filas = _leer_hoja_por_alias(ruta, ALIAS_SKU_TIENDA, CONVERTIDORES_SKU_TIENDA,
                                  REQUERIDOS_SKU_TIENDA, adv)
    if not filas:
        return 0

    presentes = {c for f in filas for c in f}
    columnas = [c for c in COLUMNAS_BD_SKU_TIENDA if c in presentes]

    ausentes = [c for c in COLUMNAS_BD_SKU_TIENDA if c not in presentes]
    if ausentes:
        adv.append(f"{ruta.name}: el archivo no trae {sorted(ausentes)}. Esas "
                    f"columnas se dejan como estaban; sólo se actualizan las "
                    f"{len(columnas)} que sí vienen.")

    total = 0
    for i in range(0, len(filas), TAMANO_LOTE):
        total += upsert_lote(cur, "catalogo_sku_tienda", columnas, ["sku", "tienda"],
                              filas[i:i + TAMANO_LOTE], adv)
    return total


def main() -> int:
    ap = argparse.ArgumentParser(description="Carga sucursales y catálogo de SKU por tienda a Postgres.")
    ap.add_argument("--sucursales", type=Path, default=None, help="Listado_sucursales....xlsx")
    ap.add_argument("--sku", type=Path, nargs="+", default=[], help="Uno o más Catalogo_Abarrotes_<tienda>.xlsx")
    ap.add_argument("--dsn", default=None)
    args = ap.parse_args()

    if not args.sucursales and not args.sku:
        print("Nada que cargar — pasa --sucursales y/o --sku.", file=sys.stderr)
        return 1

    conn = conectar(args.dsn)
    adv: list[str] = []
    t_inicio = time.time()
    try:
        with conn, conn.cursor() as cur:
            nombres = ([args.sucursales.name] if args.sucursales else []) + [p.name for p in args.sku]
            carga_id = registrar_carga_inicio(cur, "catalogos_informativos", nombres)

        estado, filas_por_hoja = "ok", {}

        if args.sucursales:
            try:
                with conn, conn.cursor() as cur:
                    n = cargar_sucursales(cur, args.sucursales, adv)
                filas_por_hoja["SUCURSALES"] = n
                print(f"  ok  SUCURSALES               {n:>9,} filas")
            except Exception as e:
                estado = "parcial"
                adv.append(f"SUCURSALES: la carga falló — {e}")
                print(f"  !!  SUCURSALES FALLÓ: {e}", file=sys.stderr)

        total_sku = 0
        for ruta in args.sku:
            try:
                with conn, conn.cursor() as cur:
                    n = cargar_catalogo_sku_tienda(cur, ruta, adv)
                total_sku += n
                print(f"  ok  {ruta.name:<40} {n:>9,} filas")
            except Exception as e:
                estado = "parcial"
                adv.append(f"{ruta.name}: la carga falló y se saltó — {e}")
                print(f"  !!  {ruta.name} FALLÓ: {e}", file=sys.stderr)
        if args.sku:
            filas_por_hoja["CATALOGO_SKU_TIENDA"] = total_sku

        with conn, conn.cursor() as cur:
            registrar_carga_fin(cur, carga_id, estado, filas_por_hoja, adv,
                                 validado=False, forzado=False, duracion_s=time.time() - t_inicio)
        return 0 if estado == "ok" else 1
    finally:
        conn.close()


if __name__ == "__main__":
    load_dotenv()
    raise SystemExit(main())
